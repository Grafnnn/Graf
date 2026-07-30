#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build SH550 facade executive-documentation albums from the working registry.

The script:
- downloads both the original and working registries from Google Drive;
- downloads and inspects the RAR archive with the extracted facade documentation;
- downloads each document referenced by the working registry;
- performs additional checks for all rows marked "ПРОВЕРИТЬ" / "НЕТ В АРХИВЕ";
- creates one ordered PDF per registry tab;
- inserts an explanatory placeholder before a partial candidate, or instead of a missing file;
- creates component ZIP archives and a final delivery ZIP;
- creates an updated Excel registry and a PDF/CSV/JSON QC report.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import os
import re
import shutil
import subprocess
import sys
import textwrap
import time
import traceback
import zipfile
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path
from typing import Iterable, Optional
from xml.sax.saxutils import escape

import fitz  # PyMuPDF
import gdown
import openpyxl
import pytesseract
import requests
from PIL import Image
from pypdf import PdfReader, PdfWriter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

WORKING_REGISTRY_ID = "1r6lNVkvetkvNAvdQVUyFpa8PC3Nb-jcD"
ORIGINAL_REGISTRY_ID = "1JCk8rtTP15rEGMZLOGcQpuQ-7MQH7WO1"
RAR_ARCHIVE_ID = "1EyHU0ZISbDJ4jgUTktAhkcNByl8w8-ii"
SCAN_AK_ID = "1Q_6hTMAUUdrSu8BZA4pCIH2-EWv6ktbm"

ALBUMS = [
    ("кронштейны", "SH550_01_Brackets.pdf", "Альбом 01. Кронштейны"),
    ("утепление 1 слой", "SH550_02_Insulation_Layer_1.pdf", "Альбом 02. Утепление — 1-й слой"),
    ("утепление 2 слой", "SH550_03_Insulation_Layer_2.pdf", "Альбом 03. Утепление — 2-й слой"),
    ("направляющие", "SH550_04_Guides_Subsystem.pdf", "Альбом 04. Направляющие и подсистема НВФ"),
    ("керамогранит", "SH550_05_Porcelain_Stoneware.pdf", "Альбом 05. Керамогранит"),
]

EXPECTED_COUNTS = {
    "кронштейны": 32,
    "утепление 1 слой": 30,
    "утепление 2 слой": 30,
    "направляющие": 39,
    "керамогранит": 64,
}

FONT_REGULAR = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
FONT_BOLD = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
PDF_FONT = "DejaVu"
PDF_FONT_BOLD = "DejaVu-Bold"


def log(msg: str) -> None:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


def register_fonts() -> None:
    if not Path(FONT_REGULAR).exists() or not Path(FONT_BOLD).exists():
        raise FileNotFoundError("DejaVu fonts are not installed")
    pdfmetrics.registerFont(TTFont(PDF_FONT, FONT_REGULAR))
    pdfmetrics.registerFont(TTFont(PDF_FONT_BOLD, FONT_BOLD))


def safe_name(value: str, max_len: int = 120) -> str:
    value = re.sub(r"[\\/:*?\"<>|\r\n]+", "_", str(value or ""))
    value = re.sub(r"\s+", " ", value).strip(" ._")
    if not value:
        value = "document"
    return value[:max_len]


def normalize(value: str) -> str:
    value = str(value or "").lower().replace("ё", "е")
    value = value.replace("№", " n ")
    value = re.sub(r"[^0-9a-zа-я]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def compact_normalize(value: str) -> str:
    return re.sub(r"[^0-9a-zа-я]+", "", str(value or "").lower().replace("ё", "е"))


def extract_drive_ids(value: str) -> list[str]:
    ids: list[str] = []
    for part in re.split(r"\s*;\s*", str(value or "")):
        if not part:
            continue
        m = re.search(r"/d/([A-Za-z0-9_-]+)", part)
        if not m:
            m = re.search(r"[?&]id=([A-Za-z0-9_-]+)", part)
        if m:
            ids.append(m.group(1))
    return ids


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def validate_pdf(path: Path) -> tuple[bool, str, int]:
    try:
        reader = PdfReader(str(path), strict=False)
        if reader.is_encrypted:
            try:
                reader.decrypt("")
            except Exception:
                return False, "PDF зашифрован и не открывается без пароля", 0
        pages = len(reader.pages)
        if pages < 1:
            return False, "PDF не содержит страниц", 0
        _ = reader.pages[0].mediabox
        return True, "", pages
    except Exception as exc:
        return False, f"{type(exc).__name__}: {exc}", 0


def download_drive_file(file_id: str, dest: Path, attempts: int = 3) -> tuple[bool, str]:
    dest.parent.mkdir(parents=True, exist_ok=True)
    if dest.exists() and dest.stat().st_size > 100:
        return True, "cache"
    tmp = dest.with_suffix(dest.suffix + ".part")
    for attempt in range(1, attempts + 1):
        try:
            tmp.unlink(missing_ok=True)
            result = gdown.download(id=file_id, output=str(tmp), quiet=True, fuzzy=True)
            if result and tmp.exists() and tmp.stat().st_size > 100:
                tmp.replace(dest)
                return True, f"gdown attempt {attempt}"
        except Exception as exc:
            last = f"gdown {type(exc).__name__}: {exc}"
        try:
            url = f"https://drive.usercontent.google.com/download?id={file_id}&export=download&confirm=t"
            with requests.get(url, timeout=180, stream=True, allow_redirects=True) as response:
                response.raise_for_status()
                with tmp.open("wb") as f:
                    for chunk in response.iter_content(1024 * 1024):
                        if chunk:
                            f.write(chunk)
            if tmp.exists() and tmp.stat().st_size > 100:
                head = tmp.read_bytes()[:200].lower()
                if b"<html" not in head and b"<!doctype html" not in head:
                    tmp.replace(dest)
                    return True, f"requests attempt {attempt}"
                last = "Google Drive returned an HTML page instead of the file"
        except Exception as exc:
            last = f"requests {type(exc).__name__}: {exc}"
        time.sleep(attempt * 2)
    tmp.unlink(missing_ok=True)
    return False, locals().get("last", "unknown download error")


def extract_text_pdf(path: Path, max_pages: int = 3) -> str:
    texts: list[str] = []
    try:
        reader = PdfReader(str(path), strict=False)
        if reader.is_encrypted:
            try:
                reader.decrypt("")
            except Exception:
                return ""
        for page in reader.pages[:max_pages]:
            try:
                texts.append(page.extract_text() or "")
            except Exception:
                pass
    except Exception:
        return ""
    return "\n".join(texts)


def ocr_pdf_pages(path: Path, page_indices: Iterable[int], dpi: int = 180) -> dict[int, str]:
    result: dict[int, str] = {}
    try:
        doc = fitz.open(str(path))
    except Exception:
        return result
    matrix = fitz.Matrix(dpi / 72.0, dpi / 72.0)
    for idx in page_indices:
        if idx < 0 or idx >= len(doc):
            continue
        try:
            pix = doc[idx].get_pixmap(matrix=matrix, alpha=False)
            image = Image.open(io.BytesIO(pix.tobytes("png")))
            text = pytesseract.image_to_string(image, lang="rus+eng", config="--psm 6")
            result[idx] = text or ""
        except Exception as exc:
            result[idx] = f"[OCR ERROR: {type(exc).__name__}: {exc}]"
    doc.close()
    return result


def verify_nodes(path: Path) -> tuple[bool, str]:
    text = extract_text_pdf(path, max_pages=10)
    if "узел" in normalize(text):
        return True, "Текстовый слой содержит обозначение узлов"
    ok, _, pages = validate_pdf(path)
    if not ok:
        return False, "PDF не прочитан"
    ocr = ocr_pdf_pages(path, range(min(pages, 4)), dpi=160)
    combined = normalize("\n".join(ocr.values()))
    if "узел" in combined:
        return True, "OCR выявил обозначение узлов на листе"
    return False, "Отдельное обозначение узлов машинной проверкой не подтверждено"


def inspect_ak_scan(path: Path) -> dict[str, dict]:
    ok, err, pages = validate_pdf(path)
    result = {
        "АК-001": {"verified": False, "pages": [], "reason": err},
        "АК-002": {"verified": False, "pages": [], "reason": err},
    }
    if not ok:
        return result
    ocr = ocr_pdf_pages(path, range(pages), dpi=220)
    compact_pages = {idx: compact_normalize(text) for idx, text in ocr.items()}
    for code in ("АК-001", "АК-002"):
        needle = compact_normalize(code)
        hits = [idx for idx, text in compact_pages.items() if needle in text]
        if hits:
            selected: list[int] = []
            for idx in hits:
                selected.append(idx)
                if idx + 1 < pages:
                    other = "ак002" if code == "АК-001" else "ак001"
                    if other not in compact_pages.get(idx + 1, ""):
                        selected.append(idx + 1)
            selected = sorted(set(selected))
            date_ok = any("18022025" in compact_pages.get(idx, "") for idx in selected)
            result[code] = {
                "verified": True,
                "pages": [idx + 1 for idx in selected],
                "reason": f"OCR обнаружил {code}" + (" и дату 18.02.2025" if date_ok else ""),
                "ocr_excerpt": " | ".join(re.sub(r"\s+", " ", ocr[i])[:220] for i in hits),
            }
    # Controlled fallback: the archive scan is four pages and previous structure suggests two 2-page protocols.
    if pages == 4:
        if not result["АК-001"]["verified"]:
            result["АК-001"].update({"pages": [1, 2], "reason": "OCR не подтвердил номер; приложены страницы 1–2 scan_0264 как неподтвержденный кандидат"})
        if not result["АК-002"]["verified"]:
            result["АК-002"].update({"pages": [3, 4], "reason": "OCR не подтвердил номер; приложены страницы 3–4 scan_0264 как неподтвержденный кандидат"})
    return result


def unpack_rar(rar_path: Path, output_dir: Path) -> tuple[bool, str, list[Path]]:
    output_dir.mkdir(parents=True, exist_ok=True)
    try:
        completed = subprocess.run(
            ["unar", "-f", "-o", str(output_dir), str(rar_path)],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            timeout=600,
            check=False,
        )
        files = [p for p in output_dir.rglob("*") if p.is_file()]
        ok = completed.returncode == 0 and bool(files)
        return ok, completed.stdout[-6000:], files
    except Exception as exc:
        return False, f"{type(exc).__name__}: {exc}", []


def archive_index(files: Iterable[Path]) -> dict[str, list[Path]]:
    idx: dict[str, list[Path]] = {}
    for path in files:
        idx.setdefault(normalize(path.name), []).append(path)
    return idx


def find_exact_archive_file(index: dict[str, list[Path]], expected_name: str) -> Optional[Path]:
    key = normalize(expected_name)
    if key in index:
        return index[key][0]
    stem = normalize(Path(expected_name).stem)
    candidates: list[tuple[float, Path]] = []
    for name, paths in index.items():
        score = SequenceMatcher(None, stem, normalize(Path(name).stem)).ratio()
        if score >= 0.94:
            candidates.append((score, paths[0]))
    if candidates:
        candidates.sort(key=lambda x: x[0], reverse=True)
        return candidates[0][1]
    return None


def find_archive_candidates(files: Iterable[Path], description: str, candidate_name: str, limit: int = 8) -> list[str]:
    target = normalize(f"{description} {candidate_name}")
    target_words = set(target.split())
    scored: list[tuple[float, Path]] = []
    for path in files:
        if path.suffix.lower() != ".pdf":
            continue
        name = normalize(path.stem)
        words = set(name.split())
        overlap = len(target_words & words) / max(1, len(target_words | words))
        ratio = SequenceMatcher(None, target, name).ratio()
        score = overlap * 0.65 + ratio * 0.35
        if score >= 0.24:
            scored.append((score, path))
    scored.sort(key=lambda x: x[0], reverse=True)
    return [f"{score:.3f}: {path.name}" for score, path in scored[:limit]]


@dataclass
class SourceSpec:
    path: str
    label: str
    pages: list[int] = field(default_factory=list)  # 1-based explicit pages; empty means all pages
    drive_id: str = ""
    sha256: str = ""
    page_count: int = 0


@dataclass
class RowResult:
    album_sheet: str
    number: int
    excel_row: int
    registry_name: str
    doc_number: str
    organization: str
    original_status: str
    found_name: str
    folder: str
    drive_url: str
    original_note: str
    final_status: str = ""
    placeholder: bool = False
    placeholder_heading: str = ""
    final_note: str = ""
    sources: list[SourceSpec] = field(default_factory=list)
    archive_candidates: list[str] = field(default_factory=list)
    album_page_start: int = 0
    album_page_end: int = 0
    component_files: list[str] = field(default_factory=list)


def row_from_cells(sheet: str, values: list) -> Optional[RowResult]:
    if not values or values[0] in (None, ""):
        return None
    try:
        number = int(float(values[0]))
    except Exception:
        return None
    padded = list(values) + [None] * (10 - len(values))
    try:
        excel_row = int(float(padded[1])) if padded[1] not in (None, "") else 0
    except Exception:
        excel_row = 0
    return RowResult(
        album_sheet=sheet,
        number=number,
        excel_row=excel_row,
        registry_name=str(padded[2] or ""),
        doc_number=str(padded[3] or ""),
        organization=str(padded[4] or ""),
        original_status=str(padded[5] or ""),
        found_name=str(padded[6] or ""),
        folder=str(padded[7] or ""),
        drive_url=str(padded[8] or ""),
        original_note=str(padded[9] or ""),
    )


def load_working_registry(path: Path) -> dict[str, list[RowResult]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    result: dict[str, list[RowResult]] = {}
    for sheet, _, _ in ALBUMS:
        if sheet not in wb.sheetnames:
            raise KeyError(f"Sheet '{sheet}' not found; sheets={wb.sheetnames}")
        ws = wb[sheet]
        rows: list[RowResult] = []
        for values in ws.iter_rows(min_row=2, max_col=10, values_only=True):
            item = row_from_cells(sheet, list(values))
            if item:
                rows.append(item)
        rows.sort(key=lambda r: r.number)
        if len(rows) != EXPECTED_COUNTS[sheet]:
            raise ValueError(f"Sheet {sheet}: expected {EXPECTED_COUNTS[sheet]} rows, got {len(rows)}")
        result[sheet] = rows
    return result


def compare_original_registry(path: Path) -> dict:
    report: dict = {"downloaded": path.exists(), "sheets": {}}
    if not path.exists():
        return report
    wb = openpyxl.load_workbook(path, data_only=True)
    report["workbook_sheets"] = wb.sheetnames
    target_names = {
        "кронштейны": "Реестр на кронштейны",
        "утепление 1 слой": "Реестр на утепление 1 слой",
        "утепление 2 слой": "Реестр на утепление 2 слой",
        "направляющие": "Реестр на направляющие",
        "керамогранит": "Реестр на керамогранит",
    }
    for short, full in target_names.items():
        if full not in wb.sheetnames:
            report["sheets"][short] = {"found": False}
            continue
        ws = wb[full]
        numbers: set[int] = set()
        for row in ws.iter_rows(values_only=True):
            for cell in row[:3]:
                try:
                    n = int(float(cell))
                except Exception:
                    continue
                if 1 <= n <= 200:
                    numbers.add(n)
                    break
        expected = EXPECTED_COUNTS[short]
        contiguous = set(range(1, expected + 1)).issubset(numbers)
        report["sheets"][short] = {
            "found": True,
            "max_row": ws.max_row,
            "detected_position_numbers": sorted(numbers),
            "expected_count": expected,
            "sequence_1_to_expected_detected": contiguous,
        }
    return report


def split_names(value: str) -> list[str]:
    return [p.strip() for p in re.split(r"\s*;\s*", str(value or "")) if p.strip()]


def process_row(
    row: RowResult,
    cache_dir: Path,
    archive_files: list[Path],
    archive_idx: dict[str, list[Path]],
    ak_scan_path: Optional[Path],
    ak_info: dict[str, dict],
) -> RowResult:
    flagged = "ПРОВЕРИТЬ" in row.original_status.upper() or "НЕТ В АРХИВЕ" in row.original_status.upper()
    row.archive_candidates = find_archive_candidates(archive_files, row.registry_name, row.found_name) if flagged else []

    # Special handling for AK-001 / AK-002 found inside scan_0264.pdf.
    code_match = re.search(r"АК-00[12]", f"{row.registry_name} {row.doc_number}", flags=re.I)
    if code_match and ak_scan_path and ak_scan_path.exists():
        code = code_match.group(0).upper()
        info = ak_info.get(code, {})
        pages = list(info.get("pages") or [])
        if pages:
            valid, err, page_count = validate_pdf(ak_scan_path)
            if valid:
                row.sources.append(SourceSpec(
                    path=str(ak_scan_path),
                    label=f"scan_0264.pdf — {code}, страницы {', '.join(map(str, pages))}",
                    pages=pages,
                    drive_id=SCAN_AK_ID,
                    sha256=sha256(ak_scan_path),
                    page_count=len(pages),
                ))
                if info.get("verified"):
                    row.final_status = "НАЙДЕНО И OCR-ПОДТВЕРЖДЕНО"
                    row.placeholder = False
                    row.final_note = str(info.get("reason") or "")
                else:
                    row.final_status = "КАНДИДАТ ПОДШИТ / НОМЕР НЕ ПОДТВЕРЖДЕН"
                    row.placeholder = True
                    row.placeholder_heading = "ТОЧНЫЙ ЛИСТ НЕ ПОДТВЕРЖДЕН"
                    row.final_note = str(info.get("reason") or "")
                return row
            row.final_note = err

    ids = extract_drive_ids(row.drive_url)
    expected_names = split_names(row.found_name)
    sources: list[SourceSpec] = []
    errors: list[str] = []
    source_count = max(len(ids), len(expected_names))
    for index in range(source_count):
        file_id = ids[index] if index < len(ids) else ""
        expected_name = expected_names[index] if index < len(expected_names) else f"source_{index + 1}.pdf"
        dest = cache_dir / f"{file_id or 'archive'}_{safe_name(expected_name, 90)}"
        if dest.suffix.lower() != ".pdf":
            dest = dest.with_suffix(".pdf")
        obtained = False
        obtain_note = ""
        if file_id:
            obtained, obtain_note = download_drive_file(file_id, dest)
            if obtained:
                valid, err, pages = validate_pdf(dest)
                if not valid:
                    errors.append(f"{expected_name}: загружен, но не является читаемым PDF: {err}")
                    obtained = False
                else:
                    sources.append(SourceSpec(
                        path=str(dest),
                        label=expected_name,
                        drive_id=file_id,
                        sha256=sha256(dest),
                        page_count=pages,
                    ))
        if not obtained:
            archive_path = find_exact_archive_file(archive_idx, expected_name)
            if archive_path:
                valid, err, pages = validate_pdf(archive_path)
                if valid:
                    sources.append(SourceSpec(
                        path=str(archive_path),
                        label=f"{expected_name} [из RAR-архива]",
                        drive_id=file_id,
                        sha256=sha256(archive_path),
                        page_count=pages,
                    ))
                    obtained = True
                else:
                    errors.append(f"{expected_name}: файл в архиве поврежден: {err}")
            if not obtained:
                errors.append(f"{expected_name}: не получен ({obtain_note or 'нет ссылки и нет точного файла в архиве'})")

    row.sources = sources
    is_partial = "ПРОВЕРИТЬ" in row.original_status.upper()
    is_missing = "НЕТ В АРХИВЕ" in row.original_status.upper()

    # A row marked as "Узлы" may be fulfilled by a combined sheet containing the nodes.
    node_verified = False
    node_reason = ""
    if is_partial and "узл" in normalize(row.registry_name) and sources:
        checks = [verify_nodes(Path(src.path)) for src in sources]
        node_verified = any(ok for ok, _ in checks)
        node_reason = "; ".join(reason for _, reason in checks)

    if not sources:
        row.final_status = "НЕ НАЙДЕНО — ВСТАВЛЕНА ЗАГЛУШКА"
        row.placeholder = True
        row.placeholder_heading = "ЛИСТ / ДОКУМЕНТ НЕ НАЙДЕН"
        row.final_note = "; ".join(filter(None, [row.original_note, *errors]))
    elif is_partial and node_verified:
        row.final_status = "НАЙДЕНО, НАЛИЧИЕ УЗЛОВ ПОДТВЕРЖДЕНО"
        row.placeholder = False
        row.final_note = "; ".join(filter(None, [node_reason, *errors]))
    elif is_partial:
        row.final_status = "ЧАСТИЧНОЕ СООТВЕТСТВИЕ — ЗАГЛУШКА + КАНДИДАТ"
        row.placeholder = True
        row.placeholder_heading = "ТОЧНОЕ СООТВЕТСТВИЕ РЕЕСТРУ НЕ НАЙДЕНО"
        extra = "Ближайшие совпадения в полном архивном поиске: " + " | ".join(row.archive_candidates[:5]) if row.archive_candidates else ""
        row.final_note = "; ".join(filter(None, [row.original_note, node_reason, extra, *errors]))
    elif is_missing:
        row.final_status = "НАЙДЕН КАНДИДАТ — ЗАГЛУШКА + КАНДИДАТ"
        row.placeholder = True
        row.placeholder_heading = "ТОЧНЫЙ ДОКУМЕНТ ИЗ РЕЕСТРА НЕ ПОДТВЕРЖДЕН"
        row.final_note = "; ".join(filter(None, [row.original_note, *errors]))
    else:
        row.final_status = "НАЙДЕНО И ПОДШИТО"
        row.placeholder = False
        row.final_note = "; ".join(errors)
    return row


def para(text: str, style: ParagraphStyle) -> Paragraph:
    return Paragraph(escape(str(text or "")).replace("\n", "<br/>"), style)


def make_cover_pdf(album_title: str, rows: list[RowResult], output: Path) -> None:
    width, height = A4
    c = canvas.Canvas(str(output), pagesize=A4)
    c.setTitle(album_title)
    c.setFont(PDF_FONT_BOLD, 16)
    c.drawCentredString(width / 2, height - 34 * mm, "ИСПОЛНИТЕЛЬНАЯ ДОКУМЕНТАЦИЯ")
    c.setFont(PDF_FONT_BOLD, 19)
    y = height - 56 * mm
    for line in textwrap.wrap(album_title, width=48):
        c.drawCentredString(width / 2, y, line)
        y -= 9 * mm
    c.setFont(PDF_FONT, 10)
    c.drawCentredString(width / 2, y - 8 * mm, "Объект: школа на 550 мест, Ново-Переделкино")
    c.drawCentredString(width / 2, y - 15 * mm, "Сборка выполнена по загруженному Excel-реестру")
    counts = {
        "Всего позиций": len(rows),
        "Найдено без заглушки": sum(1 for r in rows if not r.placeholder),
        "Позиции с информационной заглушкой": sum(1 for r in rows if r.placeholder),
        "Полностью без найденного исходного PDF": sum(1 for r in rows if not r.sources),
    }
    y2 = y - 38 * mm
    c.setFont(PDF_FONT_BOLD, 11)
    c.drawString(28 * mm, y2, "Контроль комплектности:")
    c.setFont(PDF_FONT, 10)
    for key, value in counts.items():
        y2 -= 8 * mm
        c.drawString(34 * mm, y2, f"{key}: {value}")
    c.setFont(PDF_FONT, 8.5)
    c.drawString(28 * mm, 28 * mm, f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    c.drawRightString(width - 28 * mm, 28 * mm, "Проверочная сборка ПТО")
    c.showPage()
    c.save()


def make_registry_pdf(album_title: str, rows: list[RowResult], output: Path) -> None:
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title_ru", parent=styles["Title"], fontName=PDF_FONT_BOLD, fontSize=14,
        leading=17, alignment=TA_CENTER, spaceAfter=8 * mm,
    )
    cell = ParagraphStyle(
        "cell_ru", parent=styles["BodyText"], fontName=PDF_FONT, fontSize=6.2,
        leading=7.5, alignment=TA_LEFT,
    )
    cell_bold = ParagraphStyle(
        "cell_bold_ru", parent=cell, fontName=PDF_FONT_BOLD, alignment=TA_CENTER,
    )
    doc = SimpleDocTemplate(
        str(output), pagesize=A4, rightMargin=10 * mm, leftMargin=10 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title=f"Реестр — {album_title}",
    )
    story = [para(f"Реестр состава — {album_title}", title_style)]
    data = [[
        para("№", cell_bold),
        para("Наименование по реестру", cell_bold),
        para("Номер / дата", cell_bold),
        para("Организация", cell_bold),
        para("Итог проверки", cell_bold),
    ]]
    for row in rows:
        data.append([
            para(str(row.number), cell),
            para(row.registry_name, cell),
            para(row.doc_number, cell),
            para(row.organization, cell),
            para(row.final_status, cell),
        ])
    table = Table(data, colWidths=[10 * mm, 77 * mm, 39 * mm, 31 * mm, 33 * mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9E2F3")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1F1F1F")),
        ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#888888")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 2.5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2.5),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
    ]))
    for idx, row in enumerate(rows, start=1):
        if row.placeholder:
            table.setStyle(TableStyle([("BACKGROUND", (4, idx), (4, idx), colors.HexColor("#FCE4D6"))]))
    story.append(table)
    doc.build(story)


def draw_wrapped(c: canvas.Canvas, text: str, x: float, y: float, width_chars: int, leading: float, font: str, size: float) -> float:
    c.setFont(font, size)
    for paragraph_text in str(text or "").splitlines() or [""]:
        lines = textwrap.wrap(paragraph_text, width=width_chars, break_long_words=False, replace_whitespace=False) or [""]
        for line in lines:
            c.drawString(x, y, line)
            y -= leading
        y -= leading * 0.25
    return y


def make_placeholder_pdf(row: RowResult, output: Path) -> None:
    width, height = A4
    c = canvas.Canvas(str(output), pagesize=A4)
    c.setTitle(f"Заглушка — позиция {row.number}")
    c.setStrokeColor(colors.HexColor("#C00000"))
    c.setLineWidth(2)
    c.rect(14 * mm, 14 * mm, width - 28 * mm, height - 28 * mm)
    c.setFillColor(colors.HexColor("#C00000"))
    c.setFont(PDF_FONT_BOLD, 16)
    c.drawCentredString(width / 2, height - 30 * mm, row.placeholder_heading or "ДОКУМЕНТ НЕ НАЙДЕН")
    c.setFillColor(colors.black)
    y = height - 47 * mm
    labels = [
        ("Альбом / вкладка", row.album_sheet),
        ("Позиция реестра", str(row.number)),
        ("Строка исходного Excel", str(row.excel_row)),
        ("Требуемое наименование", row.registry_name),
        ("Номер / дата", row.doc_number),
        ("Организация", row.organization),
        ("Исходный статус рабочего реестра", row.original_status),
        ("Найденный ближайший файл", row.found_name),
        ("Место хранения", row.folder),
        ("Результат повторной проверки", row.final_status),
        ("Пояснение", row.final_note),
    ]
    for label, value in labels:
        c.setFont(PDF_FONT_BOLD, 9)
        c.drawString(24 * mm, y, f"{label}:")
        y -= 5 * mm
        y = draw_wrapped(c, value or "—", 28 * mm, y, 90, 4.3 * mm, PDF_FONT, 8.5)
        y -= 2 * mm
        if y < 31 * mm:
            c.showPage()
            c.setFont(PDF_FONT, 8.5)
            y = height - 25 * mm
    c.setFont(PDF_FONT, 7.5)
    c.drawString(24 * mm, 21 * mm, "Лист сформирован автоматически для фиксации отсутствия или неполного соответствия документа реестру.")
    c.showPage()
    c.save()


def reader_for(path: Path) -> PdfReader:
    reader = PdfReader(str(path), strict=False)
    if reader.is_encrypted:
        reader.decrypt("")
    return reader


def append_pdf(writer: PdfWriter, path: Path, pages: list[int] | None = None) -> int:
    reader = reader_for(path)
    added = 0
    if pages:
        for number in pages:
            idx = number - 1
            if 0 <= idx < len(reader.pages):
                writer.add_page(reader.pages[idx])
                added += 1
    else:
        for page in reader.pages:
            writer.add_page(page)
            added += 1
    return added


def copy_component(src: Path, dest: Path) -> Path:
    dest.parent.mkdir(parents=True, exist_ok=True)
    if src.resolve() != dest.resolve():
        shutil.copy2(src, dest)
    return dest


def build_album(
    sheet: str,
    album_title: str,
    filename: str,
    rows: list[RowResult],
    output_dir: Path,
    component_root: Path,
) -> tuple[Path, Path, int]:
    album_path = output_dir / filename
    component_dir = component_root / Path(filename).stem
    component_dir.mkdir(parents=True, exist_ok=True)
    cover = component_dir / "000_cover.pdf"
    registry = component_dir / "001_registry.pdf"
    make_cover_pdf(album_title, rows, cover)
    make_registry_pdf(album_title, rows, registry)

    writer = PdfWriter()
    append_pdf(writer, cover)
    append_pdf(writer, registry)
    try:
        writer.add_outline_item("Титульный лист", 0)
        writer.add_outline_item("Реестр состава", 1)
    except Exception:
        pass

    for row in rows:
        start = len(writer.pages)
        try:
            writer.add_outline_item(f"{row.number:02d}. {row.registry_name[:100]}", start)
        except Exception:
            pass
        row.component_files = []
        if row.placeholder:
            placeholder_name = f"{row.number:03d}_00_PLACEHOLDER.pdf"
            placeholder_path = component_dir / placeholder_name
            make_placeholder_pdf(row, placeholder_path)
            append_pdf(writer, placeholder_path)
            row.component_files.append(placeholder_name)
        for source_idx, source in enumerate(row.sources, start=1):
            src = Path(source.path)
            suffix_letter = chr(ord('a') + source_idx - 1) if len(row.sources) > 1 else ""
            component_name = f"{row.number:03d}_{suffix_letter}_{safe_name(source.label, 105)}"
            if not component_name.lower().endswith(".pdf"):
                component_name += ".pdf"
            component_path = copy_component(src, component_dir / component_name)
            row.component_files.append(component_name)
            try:
                append_pdf(writer, component_path, source.pages)
            except Exception as exc:
                error_row = RowResult(**{**asdict(row), "sources": []})
                error_row.placeholder = True
                error_row.placeholder_heading = "ОШИБКА ПОДШИВКИ НАЙДЕННОГО PDF"
                error_row.final_note = f"{type(exc).__name__}: {exc}"
                error_name = f"{row.number:03d}_{suffix_letter}_APPEND_ERROR.pdf"
                error_path = component_dir / error_name
                make_placeholder_pdf(error_row, error_path)
                append_pdf(writer, error_path)
                row.component_files.append(error_name)
        if len(writer.pages) == start:
            # Defensive guarantee: one registry item must always create at least one album page.
            row.placeholder = True
            row.placeholder_heading = "ПОЗИЦИЯ НЕ СФОРМИРОВАНА"
            row.final_status = "ТЕХНИЧЕСКАЯ ЗАГЛУШКА"
            row.final_note = row.final_note or "Нет доступного PDF и не была создана исходная заглушка"
            emergency = component_dir / f"{row.number:03d}_EMERGENCY_PLACEHOLDER.pdf"
            make_placeholder_pdf(row, emergency)
            append_pdf(writer, emergency)
            row.component_files.append(emergency.name)
        row.album_page_start = start + 1
        row.album_page_end = len(writer.pages)

    writer.add_metadata({
        "/Title": album_title,
        "/Subject": "Исполнительная документация по фасадам Ш550",
        "/Author": "ПТО — автоматизированная проверочная сборка",
        "/CreationDate": datetime.now().strftime("D:%Y%m%d%H%M%S"),
    })
    with album_path.open("wb") as f:
        writer.write(f)
    valid, err, page_count = validate_pdf(album_path)
    if not valid:
        raise RuntimeError(f"Album {album_path.name} failed validation: {err}")

    component_zip = output_dir / f"{Path(filename).stem}_components.zip"
    with zipfile.ZipFile(component_zip, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
        for path in sorted(component_dir.rglob("*")):
            if path.is_file():
                zf.write(path, path.relative_to(component_dir))
    return album_path, component_zip, page_count


def make_qc_pdf(all_rows: list[RowResult], album_stats: list[dict], archive_report: dict, output: Path) -> None:
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1ru", parent=styles["Heading1"], fontName=PDF_FONT_BOLD, fontSize=15, leading=18, alignment=TA_CENTER)
    h2 = ParagraphStyle("h2ru", parent=styles["Heading2"], fontName=PDF_FONT_BOLD, fontSize=11, leading=14)
    body = ParagraphStyle("bodyru", parent=styles["BodyText"], fontName=PDF_FONT, fontSize=8, leading=10)
    cell = ParagraphStyle("cellqc", parent=body, fontSize=6.5, leading=8)
    head = ParagraphStyle("headqc", parent=cell, fontName=PDF_FONT_BOLD, alignment=TA_CENTER)
    doc = SimpleDocTemplate(str(output), pagesize=A4, rightMargin=11 * mm, leftMargin=11 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
    story = [para("Отчёт контроля комплектности исполнительной документации Ш550", h1), Spacer(1, 5 * mm)]
    total_placeholders = sum(1 for r in all_rows if r.placeholder)
    total_without_sources = sum(1 for r in all_rows if not r.sources)
    story += [
        para(f"Проверено позиций реестра: {len(all_rows)}", body),
        para(f"Позиции с информационными заглушками: {total_placeholders}", body),
        para(f"Позиции без найденного исходного PDF: {total_without_sources}", body),
        para(f"RAR-архив загружен и распакован: {'да' if archive_report.get('extracted') else 'нет'}; файлов в архиве: {archive_report.get('file_count', 0)}", body),
        Spacer(1, 5 * mm),
        para("Итоги по альбомам", h2),
    ]
    summary_data = [[para("Альбом", head), para("Позиций", head), para("Страниц PDF", head), para("Заглушек", head), para("Без исходника", head)]]
    for stat in album_stats:
        summary_data.append([
            para(stat["title"], cell), para(str(stat["items"]), cell), para(str(stat["pages"]), cell),
            para(str(stat["placeholders"]), cell), para(str(stat["without_sources"]), cell),
        ])
    table = Table(summary_data, colWidths=[90 * mm, 22 * mm, 24 * mm, 23 * mm, 25 * mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9E2F3")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]))
    story += [table, PageBreak(), para("Позиции с заглушками или замечаниями", h2)]
    flagged = [r for r in all_rows if r.placeholder or "ПОДТВЕРЖДЕНО" in r.final_status]
    flag_data = [[para("Альбом / №", head), para("Требование реестра", head), para("Итог", head), para("Пояснение", head), para("Страницы", head)]]
    for row in flagged:
        flag_data.append([
            para(f"{row.album_sheet}\n№ {row.number}", cell),
            para(row.registry_name, cell),
            para(row.final_status, cell),
            para(row.final_note, cell),
            para(f"{row.album_page_start}–{row.album_page_end}", cell),
        ])
    flag_table = Table(flag_data, colWidths=[28 * mm, 59 * mm, 37 * mm, 52 * mm, 15 * mm], repeatRows=1)
    flag_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F4B183")),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 2.5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2.5),
    ]))
    story.append(flag_table)
    doc.build(story)


def update_registry_xlsx(source: Path, all_rows: list[RowResult], album_files: dict[str, str], output: Path) -> None:
    wb = openpyxl.load_workbook(source)
    by_sheet = {sheet: [] for sheet, _, _ in ALBUMS}
    for row in all_rows:
        by_sheet[row.album_sheet].append(row)
    headers = ["Финальный статус", "Файл альбома", "Страницы альбома", "Компоненты ZIP", "Контрольное примечание"]
    for sheet, rows in by_sheet.items():
        ws = wb[sheet]
        for col, header in enumerate(headers, start=11):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        row_map = {int(ws.cell(r, 1).value): r for r in range(2, ws.max_row + 1) if str(ws.cell(r, 1).value or "").strip().isdigit()}
        for item in rows:
            r = row_map.get(item.number)
            if not r:
                continue
            values = [
                item.final_status,
                album_files[sheet],
                f"{item.album_page_start}-{item.album_page_end}",
                "\n".join(item.component_files),
                item.final_note,
            ]
            for col, value in enumerate(values, start=11):
                ws.cell(r, col, value=value)
                ws.cell(r, col).alignment = openpyxl.styles.Alignment(vertical="top", wrap_text=True)
            if item.placeholder:
                ws.cell(r, 11).fill = openpyxl.styles.PatternFill("solid", fgColor="FCE4D6")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:O{ws.max_row}"
        widths = {1: 8, 2: 12, 3: 58, 4: 31, 5: 28, 6: 32, 7: 48, 8: 38, 9: 46, 10: 55, 11: 35, 12: 33, 13: 18, 14: 55, 15: 70}
        for col, width in widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 36
    # Replace/update final summary sheet.
    if "Финальная проверка" in wb.sheetnames:
        del wb["Финальная проверка"]
    ws = wb.create_sheet("Финальная проверка", 0)
    ws.append(["Альбом", "Позиций", "Найдено без заглушки", "С заглушкой", "Без исходного PDF", "Файл PDF"])
    for sheet, _, title in ALBUMS:
        rows = by_sheet[sheet]
        ws.append([
            title, len(rows), sum(1 for r in rows if not r.placeholder), sum(1 for r in rows if r.placeholder),
            sum(1 for r in rows if not r.sources), album_files[sheet],
        ])
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E78")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, width in enumerate([42, 14, 24, 18, 22, 38], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    wb.save(output)


def write_csv(all_rows: list[RowResult], output: Path) -> None:
    fields = [
        "album_sheet", "number", "excel_row", "registry_name", "doc_number", "organization",
        "original_status", "found_name", "folder", "drive_url", "original_note", "final_status",
        "placeholder", "final_note", "album_page_start", "album_page_end", "component_files",
        "sources", "archive_candidates",
    ]
    with output.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in all_rows:
            data = asdict(row)
            data["sources"] = json.dumps(data["sources"], ensure_ascii=False)
            data["component_files"] = " | ".join(row.component_files)
            data["archive_candidates"] = " | ".join(row.archive_candidates)
            writer.writerow({key: data.get(key, "") for key in fields})


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", default="deliverables/sh550_facades_final")
    parser.add_argument("--work", default=".work_sh550")
    args = parser.parse_args()

    register_fonts()
    output_dir = Path(args.output).resolve()
    work_dir = Path(args.work).resolve()
    cache_dir = work_dir / "downloads"
    component_root = work_dir / "components"
    archive_dir = work_dir / "archive_extracted"
    output_dir.mkdir(parents=True, exist_ok=True)
    work_dir.mkdir(parents=True, exist_ok=True)

    # Remove stale final deliverables but keep directory.
    for path in output_dir.iterdir():
        if path.is_file() or path.is_symlink():
            path.unlink()
        elif path.is_dir():
            shutil.rmtree(path)

    working_xlsx = work_dir / "SH550_working_registry_facades.xlsx"
    original_xlsx = work_dir / "SH550_original_registry.xlsx"
    rar_path = work_dir / "M082-2322-09-NFS_rev1.rar"
    scan_path = cache_dir / "scan_0264.pdf"

    log("Downloading working registry")
    ok, note = download_drive_file(WORKING_REGISTRY_ID, working_xlsx)
    if not ok:
        raise RuntimeError(f"Working registry download failed: {note}")
    log(f"Working registry downloaded: {working_xlsx.stat().st_size} bytes ({note})")

    log("Downloading original registry for control comparison")
    original_ok, original_note = download_drive_file(ORIGINAL_REGISTRY_ID, original_xlsx)
    if not original_ok:
        log(f"WARNING: original registry could not be downloaded: {original_note}")

    rows_by_sheet = load_working_registry(working_xlsx)
    all_rows_initial = [row for sheet, _, _ in ALBUMS for row in rows_by_sheet[sheet]]
    log(f"Registry loaded: {len(all_rows_initial)} positions")

    original_comparison = compare_original_registry(original_xlsx) if original_ok else {"downloaded": False, "error": original_note}

    log("Downloading and extracting the RAR archive for a second-pass search")
    rar_ok, rar_note = download_drive_file(RAR_ARCHIVE_ID, rar_path, attempts=2)
    archive_files: list[Path] = []
    extraction_log = ""
    extracted = False
    if rar_ok:
        extracted, extraction_log, archive_files = unpack_rar(rar_path, archive_dir)
        log(f"RAR extraction: extracted={extracted}, files={len(archive_files)}")
    else:
        log(f"WARNING: RAR archive download failed: {rar_note}")
    idx = archive_index(archive_files)

    log("Downloading and inspecting scan_0264.pdf for AK-001 / AK-002")
    scan_ok, scan_note = download_drive_file(SCAN_AK_ID, scan_path)
    ak_info = inspect_ak_scan(scan_path) if scan_ok else {}
    log(f"AK scan status: {json.dumps(ak_info, ensure_ascii=False)[:1500]}")

    processed_by_sheet: dict[str, list[RowResult]] = {}
    for sheet, _, title in ALBUMS:
        log(f"Processing {title}: {len(rows_by_sheet[sheet])} positions")
        processed: list[RowResult] = []
        for row in rows_by_sheet[sheet]:
            try:
                processed.append(process_row(row, cache_dir, archive_files, idx, scan_path if scan_ok else None, ak_info))
            except Exception as exc:
                row.final_status = "ОШИБКА ПРОВЕРКИ — ВСТАВЛЕНА ЗАГЛУШКА"
                row.placeholder = True
                row.placeholder_heading = "ОШИБКА ОБРАБОТКИ ПОЗИЦИИ"
                row.final_note = f"{type(exc).__name__}: {exc}\n{traceback.format_exc()[-2500:]}"
                processed.append(row)
                log(f"ERROR row {sheet} #{row.number}: {row.final_note}")
        processed_by_sheet[sheet] = processed

    album_stats: list[dict] = []
    album_files: dict[str, str] = {}
    final_paths: list[Path] = []
    component_zips: list[Path] = []
    for sheet, filename, title in ALBUMS:
        log(f"Building PDF: {filename}")
        album_path, component_zip, pages = build_album(
            sheet, title, filename, processed_by_sheet[sheet], output_dir, component_root
        )
        album_files[sheet] = album_path.name
        final_paths.append(album_path)
        component_zips.append(component_zip)
        rows = processed_by_sheet[sheet]
        stat = {
            "sheet": sheet,
            "title": title,
            "file": album_path.name,
            "components_zip": component_zip.name,
            "items": len(rows),
            "pages": pages,
            "placeholders": sum(1 for r in rows if r.placeholder),
            "without_sources": sum(1 for r in rows if not r.sources),
            "size_bytes": album_path.stat().st_size,
            "sha256": sha256(album_path),
        }
        album_stats.append(stat)
        log(f"Built {album_path.name}: pages={pages}, placeholders={stat['placeholders']}, size={stat['size_bytes']}")

    all_rows = [row for sheet, _, _ in ALBUMS for row in processed_by_sheet[sheet]]
    archive_report = {
        "archive_drive_id": RAR_ARCHIVE_ID,
        "downloaded": rar_ok,
        "download_note": rar_note,
        "extracted": extracted,
        "file_count": len(archive_files),
        "pdf_count": sum(1 for p in archive_files if p.suffix.lower() == ".pdf"),
        "extraction_log_tail": extraction_log[-6000:],
    }

    qc_csv = output_dir / "SH550_QC_registry.csv"
    qc_json = output_dir / "SH550_QC_report.json"
    qc_pdf = output_dir / "SH550_QC_report.pdf"
    final_registry = output_dir / "SH550_Final_Working_Registry.xlsx"
    write_csv(all_rows, qc_csv)
    make_qc_pdf(all_rows, album_stats, archive_report, qc_pdf)
    update_registry_xlsx(working_xlsx, all_rows, album_files, final_registry)

    report = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "project": "Исполнительная документация Ш550 — фасады",
        "registry_positions": len(all_rows),
        "placeholders": sum(1 for r in all_rows if r.placeholder),
        "without_source_pdf": sum(1 for r in all_rows if not r.sources),
        "albums": album_stats,
        "archive_search": archive_report,
        "ak_scan": ak_info,
        "original_registry_comparison": original_comparison,
        "flagged_rows": [asdict(r) for r in all_rows if r.placeholder],
    }
    qc_json.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    master_zip = output_dir / "SH550_Facade_Albums_FINAL.zip"
    with zipfile.ZipFile(master_zip, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
        for path in [*final_paths, *component_zips, qc_csv, qc_json, qc_pdf, final_registry]:
            zf.write(path, path.name)
    report["master_zip"] = {
        "file": master_zip.name,
        "size_bytes": master_zip.stat().st_size,
        "sha256": sha256(master_zip),
    }
    qc_json.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    status_lines = [
        "SUCCESS",
        f"Generated: {datetime.now().isoformat()}",
        f"Registry positions: {len(all_rows)}",
        f"Placeholders: {sum(1 for r in all_rows if r.placeholder)}",
        f"Positions without source PDF: {sum(1 for r in all_rows if not r.sources)}",
        f"RAR extracted: {extracted}; archive files: {len(archive_files)}",
        "",
    ]
    for stat in album_stats:
        status_lines.append(
            f"{stat['file']}: {stat['items']} positions, {stat['pages']} pages, "
            f"{stat['placeholders']} placeholders, {stat['without_sources']} without source, "
            f"{stat['size_bytes']} bytes, sha256={stat['sha256']}"
        )
    status_lines.append(f"{master_zip.name}: {master_zip.stat().st_size} bytes, sha256={sha256(master_zip)}")
    (output_dir / "BUILD_STATUS.txt").write_text("\n".join(status_lines) + "\n", encoding="utf-8")

    log("Final structural validation")
    for path in final_paths + [qc_pdf]:
        valid, err, pages = validate_pdf(path)
        if not valid:
            raise RuntimeError(f"Final validation failed for {path.name}: {err}")
        log(f"VALID {path.name}: {pages} pages")
    log(f"DONE. Master ZIP: {master_zip} ({master_zip.stat().st_size} bytes)")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except SystemExit:
        raise
    except Exception:
        traceback.print_exc()
        raise SystemExit(1)
