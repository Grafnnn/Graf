from __future__ import annotations

import json
import re
import subprocess
import sys
from pathlib import Path

import fitz
import gdown
from openpyxl import load_workbook

OUT = Path("out")
SRC = OUT / "sources"
OUT.mkdir(exist_ok=True)
SRC.mkdir(exist_ok=True)

FILES = {
    "title.docx": "1xeh6plBTsMuZrJYYNJQUJlScqqymIurV",
    "register.pdf": "1pllLxheL6xx6vCWvIEZS3Vac9poOARTe",
    "working_registry.xlsx": "1r6lNVkvetkvNAvdQVUyFpa8PC3Nb-jcD",
    "acts.xlsx": "1R-Eiu2v45PvdMOjR_l-Vp6ienklOJ7hZ",
    "act_registries.xlsx": "1JCk8rtTP15rEGMZLOGcQpuQ-7MQH7WO1",
    "order_1.pdf": "11wHV5T4MOMO9Bl0kJodnI3etyAqlyJjf",
    "order_2.pdf": "13sOzbets8ROeKxtewGNphfNDrdvKbvgp",
    "candidate_a.pdf": "1WzUR_98y8fFUUrnQP-5GMbpsRvhEX4r9",
    "candidate_b.pdf": "1LoUfUzs8fWqaE-GNX7vuMq92353qmyFJ",
    "candidate_c.pdf": "18jHcq9NojOP3LY6zQDUC0-K3nxR6OPqA",
}


def download(name: str, file_id: str) -> Path:
    path = SRC / name
    if not path.exists() or path.stat().st_size < 100:
        print(f"Downloading {name} ({file_id})")
        result = gdown.download(id=file_id, output=str(path), quiet=False)
        if not result or not path.exists():
            raise RuntimeError(f"Could not download {name} ({file_id})")
    return path


def norm(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def inspect_xlsx(path: Path) -> dict:
    wb = load_workbook(path, data_only=False, read_only=False)
    data = {"sheets": []}
    for ws in wb.worksheets:
        strings: list[str] = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value not in (None, ""):
                    strings.append(norm(cell.value))
        joined = " | ".join(strings)
        hits = []
        for pat in [r"Фасад", r"18\.10\.2024", r"09\.11\.2024", r"16\.11\.2024", r"20\.01\.2025", r"28\.02\.2025", r"Акт", r"Реестр"]:
            if re.search(pat, joined, flags=re.I):
                hits.append(pat)
        data["sheets"].append({
            "title": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "sheet_state": ws.sheet_state,
            "print_area": str(ws.print_area),
            "orientation": ws.page_setup.orientation,
            "fit_to_width": ws.page_setup.fitToWidth,
            "fit_to_height": ws.page_setup.fitToHeight,
            "hits": hits,
            "sample": joined[:1600],
        })
    return data


def inspect_pdf(path: Path) -> dict:
    doc = fitz.open(path)
    pages = []
    for i, page in enumerate(doc):
        text = norm(page.get_text("text"))
        low = text.lower()
        score_prod = sum(1 for token in ["приказ", "производств", "ответствен", "работ"] if token in low)
        score_control = sum(1 for token in ["приказ", "строительн", "контрол", "ответствен"] if token in low)
        if score_prod >= 2 or score_control >= 3 or re.search(r"№\s*0?4|N[gоo]?\s*0?4|№\s*3|N[gоo]?\s*3", text, re.I):
            pages.append({"page": i + 1, "prod_score": score_prod, "control_score": score_control, "text": text[:3000]})
    return {"page_count": doc.page_count, "candidate_pages": pages}


def convert_office(path: Path) -> Path | None:
    outdir = OUT / "converted"
    outdir.mkdir(exist_ok=True)
    cmd = ["soffice", "--headless", "--convert-to", "pdf", "--outdir", str(outdir), str(path)]
    print("RUN", " ".join(cmd))
    proc = subprocess.run(cmd, text=True, capture_output=True)
    print(proc.stdout)
    print(proc.stderr, file=sys.stderr)
    target = outdir / (path.stem + ".pdf")
    return target if target.exists() else None


def main() -> None:
    downloaded = {name: download(name, file_id) for name, file_id in FILES.items()}
    report: dict[str, object] = {"files": {}, "xlsx": {}, "pdf": {}, "converted": {}}
    for name, path in downloaded.items():
        report["files"][name] = {"bytes": path.stat().st_size}
        if path.suffix.lower() == ".xlsx":
            report["xlsx"][name] = inspect_xlsx(path)
        elif path.suffix.lower() == ".pdf":
            report["pdf"][name] = inspect_pdf(path)
    for name in ["title.docx", "acts.xlsx", "act_registries.xlsx"]:
        converted = convert_office(downloaded[name])
        if converted:
            report["converted"][name] = inspect_pdf(converted)
    (OUT / "inspection.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
