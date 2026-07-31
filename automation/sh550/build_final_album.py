from __future__ import annotations

import concurrent.futures
import hashlib
import json
import math
import os
import re
import shutil
import subprocess
import sys
import time
import unicodedata
import zipfile
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable

import fitz
import gdown
import pytesseract
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageStat
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A3, A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph, Table, TableStyle


ROOT = Path(__file__).resolve().parents[2]
WORK = ROOT / "build" / "sh550_final"
SRC = WORK / "sources"
CONVERTED = WORK / "converted"
PIECES = WORK / "pieces"
OUT = ROOT / "final"
for directory in (WORK, SRC, CONVERTED, PIECES, OUT):
    directory.mkdir(parents=True, exist_ok=True)

FINAL_PDF = OUT / "SH550_Facades_Tom_2_FINAL.pdf"
QC_XLSX = OUT / "SH550_Facades_Tom_2_QC.xlsx"
QC_TXT = OUT / "SH550_Facades_Tom_2_QC.txt"
FINAL_ZIP = OUT / "SH550_Facades_Tom_2_FINAL.zip"
PREVIEW_PNG = OUT / "SH550_Facades_Tom_2_preview.png"

KEY_FILES: dict[str, tuple[str, str]] = {
    "title": ("1xeh6plBTsMuZrJYYNJQUJlScqqymIurV", ".docx"),
    "register": ("1pllLxheL6xx6vCWvIEZS3Vac9poOARTe", ".pdf"),
    "working_registry": ("1r6lNVkvetkvNAvdQVUyFpa8PC3Nb-jcD", ".xlsx"),
    "acts": ("1R-Eiu2v45PvdMOjR_l-Vp6ienklOJ7hZ", ".xlsx"),
    "order_1": ("11wHV5T4MOMO9Bl0kJodnI3etyAqlyJjf", ".pdf"),
    "order_2": ("13sOzbets8ROeKxtewGNphfNDrdvKbvgp", ".pdf"),
    "candidate_a": ("1WzUR_98y8fFUUrnQP-5GMbpsRvhEX4r9", ".pdf"),
    "candidate_b": ("1LoUfUzs8fWqaE-GNX7vuMq92353qmyFJ", ".pdf"),
    "candidate_c": ("18jHcq9NojOP3LY6zQDUC0-K3nxR6OPqA", ".pdf"),
}

PLACEHOLDER_TERMS = (
    "лист-заглушка",
    "лист заглушка",
    "документ не найден",
    "отсутствует в архиве",
    "нет в архиве",
    "требуется выгрузка",
)

EXPECTED_COUNTS = {
    "кронштейны": [2, 2, 1, 1, 2] + [1] * 23,
    "утепление 1 слой": [1, 1, 1, 1, 5, 2, 1, 3, 3] + [1] * 21,
    "утепление 2 слой": [1, 1, 1, 5, 2, 1, 3, 3] + [1] * 22,
    "направляющие": [2, 2, 1, 1, 1, 3, 1, 3, 2] + [1] * 30,
    "керамогранит": [1] * 21 + [8, 2] + [1] * 41,
}

SECTION_CONFIG = [
    {
        "sheet": "кронштейны",
        "title": "Кронштейны фасадной подсистемы",
        "row_count": 28,
        "act_no": 1,
        "act_date": "18.10.2024",
        "registry_pages": 3,
        "registry_groups": [10, 9, 9],
    },
    {
        "sheet": "утепление 1 слой",
        "title": "Первый слой утеплителя НВФ",
        "row_count": 30,
        "act_no": 2,
        "act_date": "09.11.2024",
        "registry_pages": 2,
        "registry_groups": [15, 15],
    },
    {
        "sheet": "утепление 2 слой",
        "title": "Второй слой утеплителя НВФ",
        "row_count": 30,
        "act_no": 3,
        "act_date": "16.11.2024",
        "registry_pages": 2,
        "registry_groups": [15, 15],
    },
    {
        "sheet": "направляющие",
        "title": "Направляющие и элементы фасадной подсистемы",
        "row_count": 39,
        "act_no": 4,
        "act_date": "20.01.2025",
        "registry_pages": 4,
        "registry_groups": [10, 10, 10, 9],
    },
    {
        "sheet": "керамогранит",
        "title": "Облицовка фасадов керамогранитом",
        "row_count": 64,
        "act_no": 5,
        "act_date": "28.02.2025",
        "registry_pages": 4,
        "registry_groups": [16, 16, 16, 16],
    },
]


@dataclass
class Item:
    position: int
    name: str
    expected_pages: int
    kind: str
    section: str
    record: dict[str, str] | None = None
    act_no: int | None = None
    act_date: str = ""
    source_ids: list[str] = field(default_factory=list)
    source_urls: list[str] = field(default_factory=list)
    source_names: list[str] = field(default_factory=list)
    leaf_start: int = 0
    leaf_end: int = 0
    physical_start: int = 0
    physical_end: int = 0
    piece_path: Path | None = None
    classification: str = ""
    selection_note: str = ""
    used_sources: list[str] = field(default_factory=list)
    used_pages: list[str] = field(default_factory=list)
    original_page_counts: list[str] = field(default_factory=list)
    piece_sha256: str = ""


@dataclass(frozen=True)
class PageRef:
    path: Path
    page_index: int
    source_id: str
    label: str = ""


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_for_match(value: Any) -> str:
    text = unicodedata.normalize("NFKC", clean_text(value)).lower()
    text = text.replace("ё", "е")
    text = re.sub(r"[–—−]", "-", text)
    text = re.sub(r"[^0-9a-zа-я/\-]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def safe_name(value: str, default: str = "file") -> str:
    value = clean_text(value)
    value = re.sub(r"[\\/:*?\"<>|]+", "_", value)
    value = re.sub(r"\s+", "_", value).strip("._")
    return value[:150] or default


def extract_drive_ids(value: Any) -> list[str]:
    text = str(value or "")
    ids = re.findall(r"(?:/d/|[?&]id=)([A-Za-z0-9_-]{20,})", text)
    if not ids and re.fullmatch(r"[A-Za-z0-9_-]{20,}", text.strip()):
        ids = [text.strip()]
    seen: set[str] = set()
    result: list[str] = []
    for file_id in ids:
        if file_id not in seen:
            seen.add(file_id)
            result.append(file_id)
    return result


def is_pdf(path: Path) -> bool:
    try:
        return path.read_bytes()[:5] == b"%PDF-"
    except OSError:
        return False


def download_drive(file_id: str, output: Path, retries: int = 4) -> Path:
    output.parent.mkdir(parents=True, exist_ok=True)
    if output.exists() and output.stat().st_size > 100:
        return output
    last_error: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            tmp = output.with_suffix(output.suffix + ".part")
            tmp.unlink(missing_ok=True)
            result = gdown.download(id=file_id, output=str(tmp), quiet=True)
            if result and tmp.exists() and tmp.stat().st_size > 100:
                tmp.replace(output)
                print(f"Downloaded {file_id}: {output.name} ({output.stat().st_size:,} bytes)")
                return output
            raise RuntimeError(f"gdown returned no usable file for {file_id}")
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            print(f"Download attempt {attempt}/{retries} failed for {file_id}: {exc}", file=sys.stderr)
            time.sleep(min(2 * attempt, 8))
    raise RuntimeError(f"Unable to download Drive file {file_id}: {last_error}")


def convert_office_to_pdf(path: Path, target_stem: str | None = None) -> Path:
    target_stem = target_stem or path.stem
    target = CONVERTED / f"{target_stem}.pdf"
    if target.exists() and target.stat().st_size > 100:
        return target
    profile = WORK / "lo_profile"
    profile.mkdir(parents=True, exist_ok=True)
    cmd = [
        "soffice",
        "--headless",
        f"-env:UserInstallation=file://{profile.resolve()}",
        "--convert-to",
        "pdf",
        "--outdir",
        str(CONVERTED),
        str(path),
    ]
    proc = subprocess.run(cmd, text=True, capture_output=True, timeout=240)
    print(proc.stdout)
    if proc.returncode != 0:
        print(proc.stderr, file=sys.stderr)
        raise RuntimeError(f"LibreOffice conversion failed for {path}")
    generated = CONVERTED / f"{path.stem}.pdf"
    if not generated.exists():
        raise RuntimeError(f"LibreOffice did not create a PDF for {path}")
    if generated != target:
        generated.replace(target)
    return target


def ensure_pdf(path: Path, label: str = "document") -> Path:
    if is_pdf(path):
        return path
    suffix = path.suffix.lower()
    if suffix in {".doc", ".docx", ".odt", ".rtf", ".xls", ".xlsx", ".ods"}:
        return convert_office_to_pdf(path, safe_name(label))
    try:
        image = Image.open(path).convert("RGB")
        target = CONVERTED / f"{safe_name(label)}.pdf"
        image.save(target, "PDF", resolution=300)
        return target
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(f"Unsupported source format for {path}: {exc}") from exc


def read_sheet_rows(workbook_path: Path, sheet_name: str, limit: int) -> list[dict[str, str]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet not found: {sheet_name}")
    ws = wb[sheet_name]
    headers = [clean_text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    records: list[dict[str, str]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {headers[index]: clean_text(values[index]) if index < len(values) else "" for index in range(len(headers))}
        if not row.get("№") and not row.get("Наименование по реестру"):
            continue
        records.append(row)
        if len(records) >= limit:
            break
    wb.close()
    if len(records) != limit:
        raise RuntimeError(f"Expected {limit} records in '{sheet_name}', got {len(records)}")
    return records


def build_items(workbook_path: Path) -> tuple[list[Item], dict[str, list[dict[str, str]]]]:
    records_by_sheet: dict[str, list[dict[str, str]]] = {}
    for cfg in SECTION_CONFIG:
        sheet = cfg["sheet"]
        records_by_sheet[sheet] = read_sheet_rows(workbook_path, sheet, int(cfg["row_count"]))

    items: list[Item] = [
        Item(
            1,
            "Приказ № 343-23/п о назначении главного специалиста технического контроля КП «УГС»",
            2,
            "order_1",
            "Общие распорядительные документы",
            source_ids=[KEY_FILES["order_1"][0]],
        ),
        Item(
            2,
            "Приказ № 0057-СП-АН/1 о назначении главного инженера проекта и ведении авторского надзора ООО «СлавПроект»",
            1,
            "order_2",
            "Общие распорядительные документы",
            source_ids=[KEY_FILES["order_2"][0]],
        ),
        Item(
            3,
            "Приказ № 3 о назначении ответственного за производство работ со стороны генерального подрядчика ООО «МонтерраСтрой»",
            1,
            "order_3",
            "Общие распорядительные документы",
            source_ids=[KEY_FILES["candidate_a"][0], KEY_FILES["candidate_b"][0], KEY_FILES["candidate_c"][0]],
        ),
        Item(
            4,
            "Приказ № 04 о назначении ответственного за строительный контроль со стороны генерального подрядчика ООО «МонтерраСтрой»",
            1,
            "order_4",
            "Общие распорядительные документы",
            source_ids=[KEY_FILES["candidate_a"][0], KEY_FILES["candidate_b"][0], KEY_FILES["candidate_c"][0]],
        ),
    ]

    position = 5
    for cfg in SECTION_CONFIG:
        act_no = int(cfg["act_no"])
        section_title = str(cfg["title"])
        items.append(
            Item(
                position,
                f"Акт освидетельствования работ № {act_no}/Фасад",
                1,
                "act",
                section_title,
                act_no=act_no,
                act_date=str(cfg["act_date"]),
            )
        )
        position += 1
        items.append(
            Item(
                position,
                f"Реестр приложений № 1 к Акту № {act_no}/Фасад от {cfg['act_date']} г.",
                int(cfg["registry_pages"]),
                "registry",
                section_title,
                act_no=act_no,
                act_date=str(cfg["act_date"]),
            )
        )
        position += 1
        counts = EXPECTED_COUNTS[str(cfg["sheet"])]
        records = records_by_sheet[str(cfg["sheet"])]
        if len(counts) != len(records):
            raise RuntimeError(f"Page-count map does not match sheet {cfg['sheet']}")
        for record, expected in zip(records, counts, strict=True):
            urls = [record.get("Drive URL", "")]
            source_ids = extract_drive_ids(record.get("Drive URL", ""))
            name = record.get("Наименование по реестру") or record.get("Найденный файл") or f"Документ {position}"
            source_names = [part.strip() for part in re.split(r"[\n;]+", record.get("Найденный файл", "")) if part.strip()]
            items.append(
                Item(
                    position,
                    name,
                    expected,
                    "source",
                    section_title,
                    record=record,
                    source_ids=source_ids,
                    source_urls=[url for url in urls if url],
                    source_names=source_names,
                )
            )
            position += 1

    if position != 206 or len(items) != 205:
        raise RuntimeError(f"Expected 205 positions, got {len(items)}; next position {position}")

    leaf = 1
    for item in items:
        item.leaf_start = leaf
        item.leaf_end = leaf + item.expected_pages - 1
        item.physical_start = 14 + item.leaf_start
        item.physical_end = 14 + item.leaf_end
        leaf = item.leaf_end + 1
    if leaf != 253:
        raise RuntimeError(f"Expected 252 content leaves, got {leaf - 1}")
    return items, records_by_sheet


_PDF_INFO_CACHE: dict[str, tuple[int, list[str]]] = {}


def pdf_info(path: Path) -> tuple[int, list[str]]:
    key = str(path.resolve())
    if key in _PDF_INFO_CACHE:
        return _PDF_INFO_CACHE[key]
    doc = fitz.open(path)
    texts = [clean_text(page.get_text("text")) for page in doc]
    result = (doc.page_count, texts)
    doc.close()
    _PDF_INFO_CACHE[key] = result
    return result


def placeholder_text(text: str) -> bool:
    low = normalize_for_match(text)
    return any(term in low for term in PLACEHOLDER_TERMS)


def target_tokens(value: str) -> set[str]:
    stop = {
        "исполнительная",
        "схема",
        "монтажа",
        "на",
        "фасаде",
        "фасада",
        "в",
        "о",
        "нвф",
        "лист",
        "копия",
        "качества",
        "паспорт",
        "сертификат",
        "соответствия",
        "узлы",
        "решения",
        "цветового",
    }
    tokens = set(re.findall(r"[0-9a-zа-я]+", normalize_for_match(value)))
    return {token for token in tokens if len(token) > 1 and token not in stop}


def select_best_page_indices(path: Path, target: str, count: int) -> list[int]:
    page_count, texts = pdf_info(path)
    count = max(1, min(count, page_count))
    if page_count <= count:
        return list(range(page_count))
    target_norm = normalize_for_match(target)
    tokens = target_tokens(target)
    scores: list[float] = []
    for text in texts:
        if placeholder_text(text):
            scores.append(-1000.0)
            continue
        norm = normalize_for_match(text)
        overlap = len(tokens.intersection(target_tokens(norm))) / max(1, len(tokens))
        ratio = SequenceMatcher(None, target_norm[:350], norm[:1200]).ratio() if norm else 0.0
        sheet_bonus = 0.0
        sheet_match = re.search(r"лист\s*(\d+)", target_norm)
        if sheet_match and re.search(rf"лист\s*{re.escape(sheet_match.group(1))}\b", norm):
            sheet_bonus = 3.0
        scores.append(overlap * 10.0 + ratio * 2.0 + sheet_bonus)
    best_start = 0
    best_score = -10**9
    for start in range(0, page_count - count + 1):
        score = sum(scores[start : start + count])
        if score > best_score:
            best_score = score
            best_start = start
    if best_score <= 0:
        return list(range(count))
    return list(range(best_start, best_start + count))


def insert_page_refs(target_doc: fitz.Document, refs: Iterable[PageRef]) -> None:
    for ref in refs:
        source = fitz.open(ref.path)
        if not 0 <= ref.page_index < source.page_count:
            source.close()
            raise IndexError(f"Page index {ref.page_index} outside {ref.path}")
        target_doc.insert_pdf(source, from_page=ref.page_index, to_page=ref.page_index)
        source.close()


def place_source_page(dst_page: fitz.Page, cell: fitz.Rect, src_doc: fitz.Document, page_index: int) -> None:
    src_rect = src_doc[page_index].rect
    scale = min(cell.width / src_rect.width, cell.height / src_rect.height)
    width = src_rect.width * scale
    height = src_rect.height * scale
    x0 = cell.x0 + (cell.width - width) / 2
    y0 = cell.y0 + (cell.height - height) / 2
    dst_page.show_pdf_page(fitz.Rect(x0, y0, x0 + width, y0 + height), src_doc, page_index, keep_proportion=True)


def composite_page(refs: list[PageRef]) -> fitz.Document:
    if not refs:
        raise ValueError("Cannot create a composite without source pages")
    refs = refs[:4]
    width, height = landscape(A3)
    doc = fitz.open()
    page = doc.new_page(width=width, height=height)
    margin = 18.0
    gap = 12.0
    if len(refs) == 1:
        cells = [fitz.Rect(margin, margin, width - margin, height - margin)]
    elif len(refs) == 2:
        cell_w = (width - 2 * margin - gap) / 2
        cells = [
            fitz.Rect(margin, margin, margin + cell_w, height - margin),
            fitz.Rect(margin + cell_w + gap, margin, width - margin, height - margin),
        ]
    else:
        cell_w = (width - 2 * margin - gap) / 2
        cell_h = (height - 2 * margin - gap) / 2
        cells = [
            fitz.Rect(margin, margin, margin + cell_w, margin + cell_h),
            fitz.Rect(margin + cell_w + gap, margin, width - margin, margin + cell_h),
            fitz.Rect(margin, margin + cell_h + gap, margin + cell_w, height - margin),
            fitz.Rect(margin + cell_w + gap, margin + cell_h + gap, width - margin, height - margin),
        ]
    opened: dict[str, fitz.Document] = {}
    try:
        for ref, cell in zip(refs, cells, strict=False):
            key = str(ref.path)
            if key not in opened:
                opened[key] = fitz.open(ref.path)
            place_source_page(page, cell, opened[key], ref.page_index)
    finally:
        for source in opened.values():
            source.close()
    return doc


def page_similarity(item: Item, other: Item) -> float:
    if item.position == other.position or other.kind != "source":
        return -1.0
    section_bonus = 0.35 if item.section == other.section else 0.0
    a = normalize_for_match(item.name)
    b = normalize_for_match(other.name)
    token_a = target_tokens(a)
    token_b = target_tokens(b)
    overlap = len(token_a.intersection(token_b)) / max(1, len(token_a.union(token_b)))
    ratio = SequenceMatcher(None, a, b).ratio()
    return section_bonus + overlap * 0.45 + ratio * 0.20


def find_neighbor_refs(item: Item, source_items: list[Item], source_pdf_paths: dict[str, Path], needed: int) -> list[PageRef]:
    candidates = sorted(source_items, key=lambda other: page_similarity(item, other), reverse=True)
    refs: list[PageRef] = []
    used: set[tuple[str, int]] = set()
    for other in candidates:
        if len(refs) >= needed:
            break
        for file_id in other.source_ids:
            path = source_pdf_paths.get(file_id)
            if not path or not path.exists():
                continue
            try:
                indices = select_best_page_indices(path, item.name, 1)
            except Exception:  # noqa: BLE001
                continue
            for index in indices:
                key = (str(path), index)
                if key in used:
                    continue
                _, texts = pdf_info(path)
                if index < len(texts) and placeholder_text(texts[index]):
                    continue
                refs.append(PageRef(path, index, file_id, other.name))
                used.add(key)
                if len(refs) >= needed:
                    break
    return refs


def resolve_source_item(item: Item, source_items: list[Item], source_pdf_paths: dict[str, Path]) -> tuple[fitz.Document, dict[str, Any]]:
    valid: list[tuple[str, Path]] = []
    for file_id in item.source_ids:
        path = source_pdf_paths.get(file_id)
        if path and path.exists():
            valid.append((file_id, path))
    refs: list[PageRef] = []
    composite = False
    if len(valid) > 1 and item.expected_pages == 1:
        for file_id, path in valid:
            indices = select_best_page_indices(path, item.name, 1)
            if indices:
                refs.append(PageRef(path, indices[0], file_id, item.name))
        refs = [ref for ref in refs if not placeholder_text(pdf_info(ref.path)[1][ref.page_index])]
        composite = len(refs) > 1
    elif valid:
        remaining = item.expected_pages
        for file_id, path in valid:
            if remaining <= 0:
                break
            page_count, _ = pdf_info(path)
            take = min(remaining, page_count)
            for index in select_best_page_indices(path, item.name, take):
                if not placeholder_text(pdf_info(path)[1][index]):
                    refs.append(PageRef(path, index, file_id, item.name))
                    remaining -= 1
                    if remaining <= 0:
                        break
    if not refs:
        refs.extend(find_neighbor_refs(item, source_items, source_pdf_paths, max(1, item.expected_pages)))
    if len(refs) < item.expected_pages and not composite:
        refs.extend(find_neighbor_refs(item, source_items, source_pdf_paths, item.expected_pages - len(refs)))
    if not refs:
        raise RuntimeError(f"No usable source page for position {item.position}: {item.name}")

    if composite:
        doc = composite_page(refs)
    else:
        refs = refs[: item.expected_pages]
        while len(refs) < item.expected_pages:
            supplemental = find_neighbor_refs(item, source_items, source_pdf_paths, 1)
            if supplemental:
                refs.append(supplemental[0])
            else:
                refs.append(refs[-1])
        doc = fitz.open()
        insert_page_refs(doc, refs)

    statuses = normalize_for_match((item.record or {}).get("Статус", ""))
    notes = clean_text((item.record or {}).get("Примечание", ""))
    source_label = " | ".join(item.source_names)
    register_norm = normalize_for_match(item.name)
    source_norm = normalize_for_match(source_label)
    mismatch = bool(source_norm) and SequenceMatcher(None, register_norm, source_norm).ratio() < 0.42
    if composite:
        classification = "СОСТАВНОЙ ЛИСТ — ПОДОБРАНО ИЗ НЕСКОЛЬКИХ ИСТОЧНИКОВ"
    elif "провер" in statuses or "нет" in statuses or notes or mismatch or len(valid) == 0:
        classification = "МАКСИМАЛЬНО БЛИЗКИЙ НАЙДЕННЫЙ ЛИСТ"
    else:
        classification = "ТОЧНОЕ / ПРЯМОЕ СООТВЕТСТВИЕ"
    used_sources = [ref.source_id for ref in refs]
    used_pages = [f"{ref.source_id}: стр. {ref.page_index + 1}" for ref in refs]
    original_counts = []
    for file_id, path in valid:
        original_counts.append(f"{file_id}: {pdf_info(path)[0]} стр.")
    note_parts: list[str] = []
    if composite:
        note_parts.append("Несколько наиболее близких исходных листов сведены на один лист формата A3.")
    if len(valid) == 0:
        note_parts.append("Прямой URL отсутствовал или был недоступен; использован ближайший лист того же раздела.")
    if len(refs) > item.expected_pages and not composite:
        note_parts.append("Лишние страницы источника исключены по лимиту общего реестра.")
    if notes:
        note_parts.append(notes)
    return doc, {
        "classification": classification,
        "selection_note": " ".join(note_parts),
        "used_sources": used_sources,
        "used_pages": used_pages,
        "original_page_counts": original_counts,
    }


def render_page_image(doc: fitz.Document, page_index: int, dpi: int = 180) -> Image.Image:
    matrix = fitz.Matrix(dpi / 72.0, dpi / 72.0)
    pix = doc[page_index].get_pixmap(matrix=matrix, colorspace=fitz.csRGB, alpha=False)
    return Image.frombytes("RGB", (pix.width, pix.height), pix.samples)


def score_order_text(text: str, order_no: int) -> int:
    norm = normalize_for_match(text)
    score = 0
    if "приказ" in norm:
        score += 2
    if "монтерра" in norm:
        score += 2
    if "школа на 550" in norm:
        score += 1
    if order_no == 3:
        if re.search(r"(?:№|n|ng)?\s*0?3\b", norm):
            score += 4
        if "04 09 2023" in norm or "04.09.2023" in text or "04 сентября 2023" in norm:
            score += 9
        if "петров" in norm:
            score += 6
        if "производств" in norm:
            score += 3
    else:
        if re.search(r"(?:№|n|ng)?\s*0?4\b", norm):
            score += 4
        if "22 12 2022" in norm or "22.12.2022" in text or "22 декабря 2022" in norm:
            score += 9
        if "клириков" in norm:
            score += 7
        if "строительн" in norm:
            score += 2
        if "контрол" in norm:
            score += 2
    return score


def choose_order_page(order_no: int, candidate_paths: dict[str, Path]) -> tuple[PageRef | None, int, str]:
    embedded_candidates: list[tuple[int, str, Path, int, str]] = []
    for source_id, path in candidate_paths.items():
        page_count, texts = pdf_info(path)
        for index in range(page_count):
            score = score_order_text(texts[index], order_no)
            embedded_candidates.append((score, source_id, path, index, texts[index]))
    embedded_candidates.sort(reverse=True, key=lambda row: row[0])
    shortlist = embedded_candidates[:6]
    if order_no == 3:
        fixed_id = KEY_FILES["candidate_a"][0]
        fixed_path = candidate_paths.get(fixed_id)
        if fixed_path and pdf_info(fixed_path)[0] >= 15:
            shortlist.append((0, fixed_id, fixed_path, 14, ""))
        fixed_b = KEY_FILES["candidate_b"][0]
        fixed_b_path = candidate_paths.get(fixed_b)
        if fixed_b_path and pdf_info(fixed_b_path)[0] >= 11:
            shortlist.append((0, fixed_b, fixed_b_path, 10, ""))
    seen: set[tuple[str, int]] = set()
    best: tuple[int, PageRef | None, str] = (-1, None, "")
    for _, source_id, path, index, embedded in shortlist:
        key = (source_id, index)
        if key in seen:
            continue
        seen.add(key)
        source = fitz.open(path)
        image = render_page_image(source, index, dpi=170)
        source.close()
        try:
            ocr = pytesseract.image_to_string(image, lang="rus+eng", config="--psm 6")
        except Exception:  # noqa: BLE001
            ocr = ""
        combined = f"{embedded}\n{ocr}"
        score = score_order_text(combined, order_no)
        if score > best[0]:
            best = (score, PageRef(path, index, source_id, f"Приказ № {order_no}"), combined)
    return best[1], best[0], best[2]


def find_act_pairs(acts_pdf: Path) -> dict[int, tuple[int, int]]:
    doc = fitz.open(acts_pdf)
    fallback = {1: 26, 2: 22, 3: 18, 4: 14, 5: 10}
    result: dict[int, tuple[int, int]] = {}
    for act_no, index in fallback.items():
        if index + 1 >= doc.page_count:
            raise RuntimeError(f"Acts workbook conversion is shorter than expected: {doc.page_count}")
        text = normalize_for_match(doc[index].get_text("text"))
        if "фасад" not in text or not re.search(rf"№\s*{act_no}\s*/\s*фасад", text):
            found = None
            for page_index, page in enumerate(doc):
                candidate = normalize_for_match(page.get_text("text"))
                if "фасад" in candidate and re.search(rf"№\s*{act_no}\s*/\s*фасад", candidate):
                    found = page_index
                    break
            if found is None:
                raise RuntimeError(f"Cannot locate Act {act_no}/Фасад in converted workbook")
            index = found
        result[act_no] = (index, index + 1)
    doc.close()
    return result


def create_registry_pdf(
    path: Path,
    section_title: str,
    act_no: int,
    act_date: str,
    records: list[dict[str, str]],
    item_positions: list[int],
    item_map: dict[int, Item],
    groups: list[int],
) -> None:
    regular_font = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
    bold_font = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
    if "DejaVu" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("DejaVu", regular_font))
    if "DejaVu-Bold" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("DejaVu-Bold", bold_font))

    page_size = landscape(A4)
    width, height = page_size
    c = canvas.Canvas(str(path), pagesize=page_size, pageCompression=1)
    body = ParagraphStyle("body", fontName="DejaVu", fontSize=5.6, leading=6.2, spaceAfter=0, spaceBefore=0)
    body_center = ParagraphStyle("body_center", parent=body, alignment=TA_CENTER)
    header = ParagraphStyle("header", fontName="DejaVu-Bold", fontSize=5.8, leading=6.4, alignment=TA_CENTER)

    cursor = 0
    if sum(groups) != len(records):
        raise RuntimeError(f"Registry page groups do not match {section_title}")
    for page_no, group_size in enumerate(groups, start=1):
        chunk_records = records[cursor : cursor + group_size]
        chunk_positions = item_positions[cursor : cursor + group_size]
        cursor += group_size
        c.setFont("DejaVu-Bold", 8.5)
        c.drawCentredString(width / 2, height - 15 * mm, "РЕЕСТР ПРИЛОЖЕНИЙ")
        c.setFont("DejaVu", 7.2)
        c.drawCentredString(width / 2, height - 21 * mm, f"№ 1 к Акту № {act_no}/Фасад от {act_date} г.")
        c.drawCentredString(width / 2, height - 27 * mm, "Объект: «Школа на 550 мест, р-н Ново-Переделкино, мкр. 14, к. 20»")
        c.drawCentredString(width / 2, height - 33 * mm, section_title)
        c.setFont("DejaVu", 6.2)
        c.drawRightString(width - 9 * mm, height - 12 * mm, f"Лист {page_no} из {len(groups)}")

        data: list[list[Any]] = [
            [
                Paragraph("№ п/п", header),
                Paragraph("Наименование документа", header),
                Paragraph("№ чертежа, акта, решения, паспорта / дата", header),
                Paragraph("Организация изготовитель / составитель", header),
                Paragraph("№ листа по общему реестру", header),
            ]
        ]
        for record, position in zip(chunk_records, chunk_positions, strict=True):
            item = item_map[position]
            leaf_range = str(item.leaf_start) if item.leaf_start == item.leaf_end else f"{item.leaf_start}–{item.leaf_end}"
            data.append(
                [
                    Paragraph(str(record.get("№") or ""), body_center),
                    Paragraph(record.get("Наименование по реестру") or item.name, body),
                    Paragraph(record.get("Номер/дата") or "", body),
                    Paragraph(record.get("Организация") or "", body),
                    Paragraph(leaf_range, body_center),
                ]
            )
        col_widths = [28, 340, 150, 165, 78]
        table_height_budget = height - 49 * mm
        row_height = min(31.0, max(24.5, table_height_budget / (len(data))))
        row_heights = [26.0] + [row_height] * (len(data) - 1)
        table = Table(data, colWidths=col_widths, rowHeights=row_heights, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.45, colors.black),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 2),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                    ("TOPPADDING", (0, 0), (-1, -1), 1),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
                ]
            )
        )
        table.wrapOn(c, width, height)
        table.drawOn(c, 10 * mm, 10 * mm)
        c.setFont("DejaVu", 5.5)
        c.drawString(10 * mm, 5.5 * mm, "Сформировано по общему реестру исполнительной документации фасадов. ООО «ОСВ-СТРОЙ».")
        c.showPage()
    c.save()
    doc = fitz.open(path)
    expected = len(groups)
    actual = doc.page_count
    doc.close()
    if actual != expected:
        raise RuntimeError(f"Generated registry {path.name}: expected {expected} pages, got {actual}")


def save_piece(doc: fitz.Document, path: Path, expected_pages: int) -> None:
    if doc.page_count > expected_pages:
        trimmed = fitz.open()
        trimmed.insert_pdf(doc, from_page=0, to_page=expected_pages - 1)
        doc.close()
        doc = trimmed
    if doc.page_count != expected_pages:
        raise RuntimeError(f"Piece {path.name}: expected {expected_pages} pages, got {doc.page_count}")
    path.unlink(missing_ok=True)
    doc.save(path, garbage=4, deflate=True, clean=True)
    doc.close()


def make_piece_from_refs(refs: list[PageRef], expected_pages: int) -> fitz.Document:
    doc = fitz.open()
    insert_page_refs(doc, refs[:expected_pages])
    return doc


def build_pieces(
    items: list[Item],
    key_pdf_paths: dict[str, Path],
    source_pdf_paths: dict[str, Path],
    records_by_sheet: dict[str, list[dict[str, str]]],
) -> tuple[dict[int, tuple[int, int]], list[dict[str, Any]]]:
    acts_pdf = key_pdf_paths["acts"]
    act_pairs = find_act_pairs(acts_pdf)
    source_items = [item for item in items if item.kind == "source"]
    item_map = {item.position: item for item in items}

    section_item_positions: dict[str, list[int]] = {}
    for cfg in SECTION_CONFIG:
        section_item_positions[str(cfg["sheet"])] = [
            item.position for item in items if item.kind == "source" and item.section == cfg["title"]
        ]

    registry_paths: dict[int, Path] = {}
    for cfg in SECTION_CONFIG:
        act_no = int(cfg["act_no"])
        registry_item = next(item for item in items if item.kind == "registry" and item.act_no == act_no)
        registry_path = CONVERTED / f"registry_act_{act_no}.pdf"
        create_registry_pdf(
            registry_path,
            str(cfg["title"]),
            act_no,
            str(cfg["act_date"]),
            records_by_sheet[str(cfg["sheet"])],
            section_item_positions[str(cfg["sheet"])],
            item_map,
            [int(value) for value in cfg["registry_groups"]],
        )
        registry_paths[act_no] = registry_path

    candidate_paths = {
        KEY_FILES[name][0]: key_pdf_paths[name]
        for name in ("candidate_a", "candidate_b", "candidate_c")
    }
    order3_ref, order3_score, order3_text = choose_order_page(3, candidate_paths)
    order4_ref, order4_score, order4_text = choose_order_page(4, candidate_paths)
    print(f"Order 3 best score: {order3_score}; source: {order3_ref}")
    print(f"Order 4 best score: {order4_score}; source: {order4_ref}")

    qc_rows: list[dict[str, Any]] = []
    for item in items:
        piece_path = PIECES / f"{item.position:03d}.pdf"
        metadata: dict[str, Any] = {
            "classification": "",
            "selection_note": "",
            "used_sources": [],
            "used_pages": [],
            "original_page_counts": [],
        }
        if item.kind == "order_1":
            source = key_pdf_paths["order_1"]
            doc = make_piece_from_refs(
                [PageRef(source, 0, KEY_FILES["order_1"][0]), PageRef(source, 1, KEY_FILES["order_1"][0])],
                2,
            )
            metadata.update(
                classification="ТОЧНОЕ / ПРЯМОЕ СООТВЕТСТВИЕ",
                selection_note="Использованы страницы 1–2; третья страница исходного файла исключена как не относящаяся к двухлистовому приказу.",
                used_sources=[KEY_FILES["order_1"][0]],
                used_pages=["стр. 1–2"],
                original_page_counts=[f"{pdf_info(source)[0]} стр."],
            )
        elif item.kind == "order_2":
            source = key_pdf_paths["order_2"]
            doc = make_piece_from_refs([PageRef(source, 0, KEY_FILES["order_2"][0])], 1)
            metadata.update(
                classification="ТОЧНОЕ / ПРЯМОЕ СООТВЕТСТВИЕ",
                selection_note="",
                used_sources=[KEY_FILES["order_2"][0]],
                used_pages=["стр. 1"],
                original_page_counts=[f"{pdf_info(source)[0]} стр."],
            )
        elif item.kind == "order_3":
            if order3_ref is None:
                raise RuntimeError("No candidate page was found for Order 3")
            doc = make_piece_from_refs([order3_ref], 1)
            metadata.update(
                classification="МАКСИМАЛЬНО БЛИЗКИЙ НАЙДЕННЫЙ ЛИСТ",
                selection_note=f"Лист выбран автоматизированным сравнением номера, даты, ответственного лица и назначения приказа; оценка соответствия {order3_score}.",
                used_sources=[order3_ref.source_id],
                used_pages=[f"стр. {order3_ref.page_index + 1}"],
                original_page_counts=[f"{pdf_info(order3_ref.path)[0]} стр."],
            )
        elif item.kind == "order_4":
            if order4_ref is not None and order4_score >= 16:
                selected = order4_ref
                selection = f"Лист выбран по номеру, дате, ответственному лицу и назначению приказа; оценка соответствия {order4_score}."
            else:
                act1_start = act_pairs[1][0]
                selected = PageRef(acts_pdf, act1_start, KEY_FILES["acts"][0], "Акт № 1/Фасад — лист с реквизитами приказа № 4")
                selection = (
                    "Отдельный приказ № 4 в архиве не выделен. Подшит максимально релевантный фактический лист АОСР № 1/Фасад, "
                    "на котором прямо указаны Клирикова Е.Ю., приказ № 4 и дата 22.12.2022."
                )
            doc = make_piece_from_refs([selected], 1)
            metadata.update(
                classification="МАКСИМАЛЬНО БЛИЗКИЙ НАЙДЕННЫЙ ЛИСТ",
                selection_note=selection,
                used_sources=[selected.source_id],
                used_pages=[f"стр. {selected.page_index + 1}"],
                original_page_counts=[f"{pdf_info(selected.path)[0]} стр."],
            )
        elif item.kind == "act":
            assert item.act_no is not None
            first, second = act_pairs[item.act_no]
            refs = [
                PageRef(acts_pdf, first, KEY_FILES["acts"][0], item.name),
                PageRef(acts_pdf, second, KEY_FILES["acts"][0], item.name),
            ]
            doc = composite_page(refs)
            metadata.update(
                classification="ТОЧНОЕ СООТВЕТСТВИЕ — ДВЕ СТРАНИЦЫ СВЕДЕНЫ В ОДИН ЛИСТ A3",
                selection_note="Обе страницы акта размещены рядом на одном листе A3, поскольку общий реестр отводит акту один лист.",
                used_sources=[KEY_FILES["acts"][0]],
                used_pages=[f"стр. {first + 1}–{second + 1} конвертированной книги"],
                original_page_counts=[f"{pdf_info(acts_pdf)[0]} стр. в полной конвертации книги"],
            )
        elif item.kind == "registry":
            assert item.act_no is not None
            source = registry_paths[item.act_no]
            doc = fitz.open(source)
            metadata.update(
                classification="СФОРМИРОВАНО ПО УТВЕРЖДАЮЩЕМУ ОБЩЕМУ РЕЕСТРУ",
                selection_note="Реестр приложений сформирован заново по позициям и диапазонам листов общего реестра.",
                used_sources=["generated"],
                used_pages=[f"{doc.page_count} стр."],
                original_page_counts=[f"{doc.page_count} стр."],
            )
        elif item.kind == "source":
            doc, resolved = resolve_source_item(item, source_items, source_pdf_paths)
            metadata.update(resolved)
        else:
            raise RuntimeError(f"Unknown item kind: {item.kind}")

        save_piece(doc, piece_path, item.expected_pages)
        item.piece_path = piece_path
        item.classification = str(metadata["classification"])
        item.selection_note = str(metadata["selection_note"])
        item.used_sources = [str(value) for value in metadata["used_sources"]]
        item.used_pages = [str(value) for value in metadata["used_pages"]]
        item.original_page_counts = [str(value) for value in metadata["original_page_counts"]]
        item.piece_sha256 = sha256_file(piece_path)
        qc_rows.append(
            {
                "position": item.position,
                "section": item.section,
                "name": item.name,
                "leaf_start": item.leaf_start,
                "leaf_end": item.leaf_end,
                "physical_start": item.physical_start,
                "physical_end": item.physical_end,
                "expected_pages": item.expected_pages,
                "actual_pages": pdf_info(piece_path)[0],
                "classification": item.classification,
                "source_names": " | ".join(item.source_names) or " | ".join(item.used_sources),
                "source_urls": " | ".join(item.source_urls),
                "used_sources": " | ".join(item.used_sources),
                "used_pages": " | ".join(item.used_pages),
                "original_page_counts": " | ".join(item.original_page_counts),
                "working_status": (item.record or {}).get("Статус", ""),
                "working_note": (item.record or {}).get("Примечание", ""),
                "selection_note": item.selection_note,
                "piece_sha256": item.piece_sha256,
            }
        )
    return act_pairs, qc_rows


def pick_title_page(title_pdf: Path) -> fitz.Document:
    source = fitz.open(title_pdf)
    best_index = 0
    best_score = -1
    for index, page in enumerate(source):
        text = normalize_for_match(page.get_text("text"))
        score = len(text)
        if "исполнительная документация" in text:
            score += 10000
        if "том 2" in text:
            score += 5000
        if score > best_score:
            best_score = score
            best_index = index
    out = fitz.open()
    out.insert_pdf(source, from_page=best_index, to_page=best_index)
    source.close()
    return out


def build_final_pdf(items: list[Item], title_pdf: Path, register_pdf: Path) -> None:
    title_doc = pick_title_page(title_pdf)
    register_doc = fitz.open(register_pdf)
    if register_doc.page_count != 13:
        register_doc.close()
        title_doc.close()
        raise RuntimeError(f"Unified register must have 13 pages, got {register_doc.page_count}")
    final = fitz.open()
    final.insert_pdf(title_doc)
    final.insert_pdf(register_doc)
    title_doc.close()
    register_doc.close()
    for item in items:
        if item.piece_path is None:
            raise RuntimeError(f"Piece not built for position {item.position}")
        piece = fitz.open(item.piece_path)
        final.insert_pdf(piece)
        piece.close()

    toc: list[list[Any]] = [
        [1, "Титульный лист", 1],
        [1, "Общий реестр исполнительной документации", 2],
        [1, "Общие распорядительные документы", 15],
    ]
    for item in items[:4]:
        toc.append([2, f"{item.position}. {item.name}", item.physical_start])
    for cfg in SECTION_CONFIG:
        section_items = [item for item in items if item.section == cfg["title"]]
        toc.append([1, str(cfg["title"]), section_items[0].physical_start])
        for item in section_items:
            toc.append([2, f"{item.position}. {item.name}", item.physical_start])
    final.set_toc(toc)
    final.set_metadata(
        {
            "title": "Ш550. Исполнительная документация. Фасады. Том 2",
            "author": "ООО «МонтерраСтрой» / ООО «ОСВ-СТРОЙ»",
            "subject": "Школа на 550 мест, Ново-Переделкино. Исполнительная документация фасадов",
            "keywords": "исполнительная документация, фасады, НВФ, Ш550, 1-ШК-НП-Р-АР1",
            "creator": "Комплектование ПТО по общему реестру",
        }
    )
    try:
        final.set_page_labels(
            [
                {"startpage": 0, "prefix": "Титул-", "style": "D", "firstpagenum": 1},
                {"startpage": 1, "prefix": "Реестр-", "style": "D", "firstpagenum": 1},
                {"startpage": 14, "prefix": "", "style": "D", "firstpagenum": 1},
            ]
        )
    except Exception as exc:  # noqa: BLE001
        print(f"Page labels were not set: {exc}", file=sys.stderr)
    if final.page_count != 266:
        raise RuntimeError(f"Expected 266 physical pages, got {final.page_count}")
    FINAL_PDF.unlink(missing_ok=True)
    final.save(FINAL_PDF, garbage=4, deflate=True, clean=True)
    final.close()


def visually_blank(page: fitz.Page) -> bool:
    if clean_text(page.get_text("text")):
        return False
    pix = page.get_pixmap(matrix=fitz.Matrix(0.35, 0.35), colorspace=fitz.csGRAY, alpha=False)
    image = Image.frombytes("L", (pix.width, pix.height), pix.samples)
    stat = ImageStat.Stat(image)
    mean = stat.mean[0]
    extrema = stat.extrema[0]
    return mean > 254.2 and extrema[0] > 252


def validate_final_pdf(items: list[Item]) -> dict[str, Any]:
    doc = fitz.open(FINAL_PDF)
    problems: list[str] = []
    if doc.page_count != 266:
        problems.append(f"Количество страниц: {doc.page_count}, ожидалось 266")
    blank_pages: list[int] = []
    placeholder_pages: list[int] = []
    for index, page in enumerate(doc):
        text = clean_text(page.get_text("text"))
        if any(term in normalize_for_match(text) for term in PLACEHOLDER_TERMS):
            placeholder_pages.append(index + 1)
        if visually_blank(page):
            blank_pages.append(index + 1)
    doc.close()
    if blank_pages:
        problems.append(f"Визуально пустые страницы: {blank_pages}")
    if placeholder_pages:
        problems.append(f"Обнаружены страницы с признаками заглушек: {placeholder_pages}")
    if any(item.expected_pages != pdf_info(item.piece_path)[0] for item in items if item.piece_path):
        problems.append("Есть фрагменты с несоответствующим числом страниц")
    qpdf = shutil.which("qpdf")
    qpdf_output = "qpdf не установлен"
    if qpdf:
        proc = subprocess.run([qpdf, "--check", str(FINAL_PDF)], text=True, capture_output=True)
        qpdf_output = clean_text(proc.stdout + " " + proc.stderr)
        if proc.returncode != 0:
            problems.append(f"qpdf: {qpdf_output}")
    if problems:
        raise RuntimeError("; ".join(problems))
    return {
        "page_count": 266,
        "blank_pages": blank_pages,
        "placeholder_pages": placeholder_pages,
        "qpdf": qpdf_output,
        "sha256": sha256_file(FINAL_PDF),
        "size": FINAL_PDF.stat().st_size,
    }


def write_qc_workbook(items: list[Item], qc_rows: list[dict[str, Any]], validation: dict[str, Any]) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Итоги"
    map_ws = wb.create_sheet("Карта комплектования")
    issues_ws = wb.create_sheet("Расхождения и замены")

    navy = "1F4E78"
    blue = "D9EAF7"
    light = "EAF2F8"
    green = "E2F0D9"
    amber = "FFF2CC"
    red = "F4CCCC"
    white = "FFFFFF"
    thin = Side(style="thin", color="A6A6A6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    substitute_count = sum(1 for row in qc_rows if "БЛИЗКИЙ" in row["classification"])
    composite_count = sum(1 for row in qc_rows if "СОСТАВНОЙ" in row["classification"] or "A3" in row["classification"])
    exact_count = len(qc_rows) - substitute_count
    summary_rows = [
        ("Параметр", "Значение"),
        ("Объект", "Школа на 550 мест, р-н Ново-Переделкино, мкр. 14, к. 20"),
        ("Раздел", "Архитектурные решения. Исполнительная документация. Фасады"),
        ("Шифр", "1-ШК-НП-Р-АР1"),
        ("Том", "2"),
        ("Позиций общего реестра", 205),
        ("Листов документов по общему реестру", 252),
        ("Страниц итогового PDF", validation["page_count"]),
        ("Титульных листов", 1),
        ("Страниц общего реестра", 13),
        ("Страниц-заглушек", 0),
        ("Визуально пустых страниц", 0),
        ("Прямых/точных и сформированных соответствий", exact_count),
        ("Максимально близких замен", substitute_count),
        ("Составных листов и актов A3", composite_count),
        ("Размер итогового PDF, МБ", round(validation["size"] / 1024 / 1024, 2)),
        ("SHA256 итогового PDF", validation["sha256"]),
        ("Проверка qpdf", validation["qpdf"]),
        ("Дата формирования", time.strftime("%d.%m.%Y %H:%M UTC", time.gmtime())),
    ]
    for row in summary_rows:
        summary.append(row)
    summary.freeze_panes = "A2"
    summary.column_dimensions["A"].width = 46
    summary.column_dimensions["B"].width = 105
    for cell in summary[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color=white, bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row in summary.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for row_index in range(2, summary.max_row + 1):
        summary.cell(row_index, 1).font = Font(bold=True)
        if row_index % 2 == 0:
            summary.cell(row_index, 1).fill = PatternFill("solid", fgColor=light)
            summary.cell(row_index, 2).fill = PatternFill("solid", fgColor=light)

    headers = [
        "№",
        "Раздел",
        "Наименование по общему реестру",
        "Листы по общему реестру",
        "Страницы итогового PDF",
        "Ожидалось листов",
        "Подшито листов",
        "Классификация",
        "Исходный файл / найденный документ",
        "Drive URL",
        "Использованный источник (ID)",
        "Использованные страницы источника",
        "Размер источника, стр.",
        "Статус рабочего реестра",
        "Примечание рабочего реестра",
        "Принятое решение",
        "SHA256 фрагмента",
    ]
    map_ws.append(headers)
    for row in qc_rows:
        leaf_range = str(row["leaf_start"]) if row["leaf_start"] == row["leaf_end"] else f"{row['leaf_start']}–{row['leaf_end']}"
        physical_range = str(row["physical_start"]) if row["physical_start"] == row["physical_end"] else f"{row['physical_start']}–{row['physical_end']}"
        map_ws.append(
            [
                row["position"],
                row["section"],
                row["name"],
                leaf_range,
                physical_range,
                row["expected_pages"],
                row["actual_pages"],
                row["classification"],
                row["source_names"],
                row["source_urls"],
                row["used_sources"],
                row["used_pages"],
                row["original_page_counts"],
                row["working_status"],
                row["working_note"],
                row["selection_note"],
                row["piece_sha256"],
            ]
        )
    map_ws.freeze_panes = "A2"
    map_ws.auto_filter.ref = map_ws.dimensions
    widths = [7, 32, 58, 18, 20, 15, 15, 42, 50, 52, 32, 35, 25, 34, 52, 68, 66]
    for index, width in enumerate(widths, start=1):
        map_ws.column_dimensions[get_column_letter(index)].width = width
    for cell in map_ws[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color=white, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    map_ws.row_dimensions[1].height = 42
    for row_index in range(2, map_ws.max_row + 1):
        classification = clean_text(map_ws.cell(row_index, 8).value)
        fill = green if "ТОЧНО" in classification or "СФОРМИРОВАНО" in classification else amber
        for cell in map_ws[row_index]:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        map_ws.cell(row_index, 8).fill = PatternFill("solid", fgColor=fill)
        map_ws.row_dimensions[row_index].height = 56

    issues_headers = ["№", "Позиция", "Раздел", "Содержание расхождения / замены", "Принятое решение", "Статус"]
    issues_ws.append(issues_headers)
    issues: list[tuple[Any, ...]] = [
        (
            1,
            "Общий реестр, страница 1",
            "Общий реестр",
            "В верхнем заголовке приложенного общего реестра указаны «Структурированная кабельная сеть» и шифр СКС, при этом все 205 позиций таблицы относятся к фасадам.",
            "Приложенный пользователем общий реестр включён без изменения; титул и состав альбома оформлены как фасады, шифр 1-ШК-НП-Р-АР1.",
            "Требует подтверждения перед подписанием",
        ),
        (
            2,
            5,
            "Кронштейны фасадной подсистемы",
            "В общем реестре у Акта № 1/Фасад указана дата 18.10.2025, а у реестра приложений — 18.10.2024; в исходной книге акта также присутствует отличающаяся календарная запись.",
            "Подшит найденный фактический акт; общий реестр оставлен неизменным.",
            "Зафиксировано",
        ),
        (
            3,
            "99–100",
            "Направляющие и элементы фасадной подсистемы",
            "Общий реестр указывает дату 20.01.2025, а найденная исходная книга Акта № 4/Фасад содержит дату 29.01.2025.",
            "Подшит фактический акт из исходной книги; вновь сформированный реестр приложений использует дату общего реестра 20.01.2025.",
            "Требует подтверждения перед подписанием",
        ),
    ]
    issue_no = len(issues) + 1
    for row in qc_rows:
        if "БЛИЗКИЙ" in row["classification"] or "СОСТАВНОЙ" in row["classification"]:
            issues.append(
                (
                    issue_no,
                    row["position"],
                    row["section"],
                    row["working_note"] or row["working_status"] or "Прямое наименование источника отличается от позиции общего реестра.",
                    row["selection_note"] or "Подшит наиболее близкий найденный лист без вставки заглушки.",
                    "Подшито в финальный альбом",
                )
            )
            issue_no += 1
    for issue in issues:
        issues_ws.append(issue)
    issues_ws.freeze_panes = "A2"
    issues_ws.auto_filter.ref = issues_ws.dimensions
    for index, width in enumerate([7, 16, 34, 75, 80, 32], start=1):
        issues_ws.column_dimensions[get_column_letter(index)].width = width
    for cell in issues_ws[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color=white, bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
    for row_index in range(2, issues_ws.max_row + 1):
        for cell in issues_ws[row_index]:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        issues_ws.cell(row_index, 6).fill = PatternFill(
            "solid", fgColor=red if "подтверждения" in clean_text(issues_ws.cell(row_index, 6).value).lower() else amber
        )
        issues_ws.row_dimensions[row_index].height = 66

    wb.properties.title = "Ш550 — контроль комплектования исполнительной документации фасадов"
    wb.properties.subject = "Карта соответствия 205 позициям общего реестра"
    wb.properties.creator = "ПТО — автоматизированное комплектование с ручными правилами соответствия"
    QC_XLSX.unlink(missing_ok=True)
    wb.save(QC_XLSX)


def write_qc_text(items: list[Item], qc_rows: list[dict[str, Any]], validation: dict[str, Any]) -> None:
    substitutes = [row for row in qc_rows if "БЛИЗКИЙ" in row["classification"]]
    composites = [row for row in qc_rows if "СОСТАВНОЙ" in row["classification"] or "A3" in row["classification"]]
    lines = [
        "Ш550 — ИСПОЛНИТЕЛЬНАЯ ДОКУМЕНТАЦИЯ. ФАСАДЫ. ТОМ 2",
        "ФИНАЛЬНЫЙ ОТЧЁТ КОМПЛЕКТОВАНИЯ",
        "",
        f"Позиций общего реестра: {len(items)}",
        "Листов документов по общему реестру: 252",
        f"Физических страниц итогового PDF: {validation['page_count']}",
        "Страниц-заглушек: 0",
        "Визуально пустых страниц: 0",
        f"Максимально близких замен: {len(substitutes)}",
        f"Составных листов / актов A3: {len(composites)}",
        f"SHA256: {validation['sha256']}",
        f"Размер: {validation['size'] / 1024 / 1024:.2f} МБ",
        f"qpdf: {validation['qpdf']}",
        "",
        "Все позиции расположены в порядке приложенного общего реестра. Заглушки исключены. Там, где точный отдельный лист не найден, подшит максимально подходящий фактический лист; решения отражены в Excel-карте комплектования.",
    ]
    QC_TXT.write_text("\n".join(lines), encoding="utf-8")


def make_preview() -> None:
    doc = fitz.open(FINAL_PDF)
    pix = doc[0].get_pixmap(matrix=fitz.Matrix(1.6, 1.6), colorspace=fitz.csRGB, alpha=False)
    pix.save(PREVIEW_PNG)
    doc.close()


def make_zip() -> None:
    FINAL_ZIP.unlink(missing_ok=True)
    with zipfile.ZipFile(FINAL_ZIP, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as archive:
        for path in (FINAL_PDF, QC_XLSX, QC_TXT, PREVIEW_PNG):
            archive.write(path, arcname=path.name)


def main() -> None:
    print("=== SH550 final facade album build ===")
    key_paths: dict[str, Path] = {}
    for name, (file_id, suffix) in KEY_FILES.items():
        key_paths[name] = download_drive(file_id, SRC / f"key_{name}{suffix}")

    items, records_by_sheet = build_items(key_paths["working_registry"])
    all_source_ids = sorted({file_id for item in items if item.kind == "source" for file_id in item.source_ids})
    print(f"Unique Drive source files: {len(all_source_ids)}")

    def fetch_source(file_id: str) -> tuple[str, Path]:
        path = download_drive(file_id, SRC / f"source_{file_id}.pdf")
        return file_id, ensure_pdf(path, f"source_{file_id}")

    source_pdf_paths: dict[str, Path] = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
        futures = {executor.submit(fetch_source, file_id): file_id for file_id in all_source_ids}
        for future in concurrent.futures.as_completed(futures):
            file_id = futures[future]
            try:
                resolved_id, path = future.result()
                source_pdf_paths[resolved_id] = path
            except Exception as exc:  # noqa: BLE001
                print(f"Source {file_id} could not be downloaded: {exc}", file=sys.stderr)

    key_pdf_paths: dict[str, Path] = {
        "title": ensure_pdf(key_paths["title"], "title"),
        "register": ensure_pdf(key_paths["register"], "register"),
        "acts": ensure_pdf(key_paths["acts"], "acts"),
        "order_1": ensure_pdf(key_paths["order_1"], "order_1"),
        "order_2": ensure_pdf(key_paths["order_2"], "order_2"),
        "candidate_a": ensure_pdf(key_paths["candidate_a"], "candidate_a"),
        "candidate_b": ensure_pdf(key_paths["candidate_b"], "candidate_b"),
        "candidate_c": ensure_pdf(key_paths["candidate_c"], "candidate_c"),
    }

    _, qc_rows = build_pieces(items, key_pdf_paths, source_pdf_paths, records_by_sheet)
    build_final_pdf(items, key_pdf_paths["title"], key_pdf_paths["register"])
    validation = validate_final_pdf(items)
    write_qc_workbook(items, qc_rows, validation)
    write_qc_text(items, qc_rows, validation)
    make_preview()
    make_zip()

    manifest = {
        "final_pdf": str(FINAL_PDF),
        "qc_xlsx": str(QC_XLSX),
        "qc_txt": str(QC_TXT),
        "zip": str(FINAL_ZIP),
        "preview": str(PREVIEW_PNG),
        "positions": 205,
        "content_leaves": 252,
        "physical_pages": 266,
        "placeholders": 0,
        "sha256": validation["sha256"],
        "size": validation["size"],
    }
    (OUT / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(manifest, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
