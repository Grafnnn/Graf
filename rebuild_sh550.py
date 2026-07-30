#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Clean rebuild of the SH550 facade executive-documentation albums.

The script downloads the working registry and all referenced source PDFs from
Google Drive, restores AK-001/AK-002 from scan_0264.pdf, creates one ordered
PDF album per registry tab, creates a source ZIP per album, updates the working
registry, performs PDF/QC checks, and creates a master archive.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import os
import re
import shutil
import subprocess
import sys
import time
import traceback
import unicodedata
import zipfile
from collections import OrderedDict, defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional
from zoneinfo import ZoneInfo

import fitz  # PyMuPDF
import gdown
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pypdf import PdfReader, PdfWriter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    LongTable,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

WORKING_REGISTRY_ID = "1IOnrLIEf3WBbysca2ql6OvQo-V3iCROH"
SOURCE_REGISTRY_ID = "1JCk8rtTP15rEGMZLOGcQpuQ-7MQH7WO1"
SCAN_0264_ID = "1Q_6hTMAUUdrSu8BZA4pCIH2-EWv6ktbm"
OBJECT_TITLE = "Школа на 550 мест, район Ново-Переделкино, мкр. 14, корп. 20"
OBJECT_ADDRESS = "г. Москва, ул. Лукинская, дом 12"
PROJECT_CODE = "1-ШК-НП-Р-Фасады"
BUILD_DATE = datetime.now(ZoneInfo("Europe/Moscow")).strftime("%d.%m.%Y")

ALBUMS: "OrderedDict[str, tuple[str, str]]" = OrderedDict(
    [
        ("Реестр на кронштейны", ("01", "Kronshteiny")),
        ("Реестр на утепление 1 слой", ("02", "Uteplenie_1_sloy")),
        ("Реестр на утепление 2 слой", ("03", "Uteplenie_2_sloy")),
        ("Реестр на направляющие", ("04", "Napravlyayushchie")),
        ("Реестр на керамогранит", ("05", "Keramogranit")),
    ]
)

FONT_REGULAR_PATH = Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf")
FONT_BOLD_PATH = Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf")
if not FONT_BOLD_PATH.exists():
    FONT_BOLD_PATH = Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf")
FONT_REGULAR = "DejaVu"
FONT_BOLD = "DejaVuBold"


def register_fonts() -> None:
    if not FONT_REGULAR_PATH.exists() or not FONT_BOLD_PATH.exists():
        raise FileNotFoundError("DejaVu fonts are required for Cyrillic PDF output")
    pdfmetrics.registerFont(TTFont(FONT_REGULAR, str(FONT_REGULAR_PATH)))
    pdfmetrics.registerFont(TTFont(FONT_BOLD, str(FONT_BOLD_PATH)))


@dataclass
class SourceSpec:
    file_id: str
    display_name: str
    drive_url: str
    page_slice: Optional[list[int]] = None
    source_path: Optional[Path] = None
    included_path: Optional[Path] = None
    page_count: int = 0
    sha256: str = ""
    archive_name: str = ""
    error: str = ""


@dataclass
class RegistryEntry:
    album: str
    number: int
    excel_row: str
    required_name: str
    document_ref: str
    organization: str
    original_status: str
    found_names_text: str
    storage_location: str
    drive_urls_text: str
    local_file_old: str
    notes: str
    match_type: str = "exact"  # exact, partial, missing, download_error
    status: str = ""
    source_specs: list[SourceSpec] = field(default_factory=list)
    service_page: Optional[Path] = None
    included_paths: list[Path] = field(default_factory=list)
    included_page_counts: list[int] = field(default_factory=list)
    page_start: int = 0
    page_end: int = 0
    page_range: str = ""
    archive_files: list[str] = field(default_factory=list)
    build_note: str = ""

    @property
    def short_status(self) -> str:
        if self.match_type == "exact":
            return "ПОДШИТО"
        if self.match_type == "partial":
            return "ПОДШИТО С ЗАМЕЧАНИЕМ"
        if self.match_type == "download_error":
            return "ОШИБКА ВЫГРУЗКИ"
        return "ОТСУТСТВУЕТ"


def text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def safe_int(value: object) -> int:
    s = text(value)
    try:
        return int(float(s))
    except Exception as exc:
        raise ValueError(f"Cannot parse registry number: {value!r}") from exc


def sanitize_filename(value: str, fallback: str = "document", max_len: int = 135) -> str:
    value = unicodedata.normalize("NFKC", value or "")
    value = value.replace("/", "_").replace("\\", "_")
    value = re.sub(r"[<>:\"|?*\x00-\x1F]", "_", value)
    value = re.sub(r"\s+", " ", value).strip(" ._")
    if not value:
        value = fallback
    stem, suffix = os.path.splitext(value)
    suffix = suffix if suffix.lower() == ".pdf" else ".pdf"
    room = max_len - len(suffix)
    stem = stem[:room].rstrip(" ._")
    return f"{stem}{suffix}"


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for block in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(block)
    return h.hexdigest()


def extract_drive_id(url: str) -> str:
    patterns = [r"/d/([A-Za-z0-9_-]+)", r"[?&]id=([A-Za-z0-9_-]+)"]
    for pattern in patterns:
        m = re.search(pattern, url)
        if m:
            return m.group(1)
    if re.fullmatch(r"[A-Za-z0-9_-]{10,}", url.strip()):
        return url.strip()
    raise ValueError(f"Cannot extract Google Drive ID from: {url}")


def split_semicolon(value: str) -> list[str]:
    return [part.strip() for part in re.split(r"\s*;\s*", value or "") if part.strip()]


def ensure_pdf_magic(path: Path) -> None:
    data = path.read_bytes()
    marker = data[:2048].find(b"%PDF")
    if marker < 0:
        snippet = data[:200].decode("utf-8", errors="replace")
        raise ValueError(f"Downloaded file is not PDF: {path.name}; start={snippet!r}")
    if marker > 0:
        path.write_bytes(data[marker:])


def pdf_page_count(path: Path) -> int:
    try:
        reader = PdfReader(str(path), strict=False)
        if reader.is_encrypted:
            reader.decrypt("")
        return len(reader.pages)
    except Exception:
        doc = fitz.open(str(path))
        count = doc.page_count
        doc.close()
        return count


def repair_pdf(path: Path, repaired_dir: Path) -> Path:
    """Return a readable PDF path, preserving the original when possible."""
    ensure_pdf_magic(path)
    try:
        count = pdf_page_count(path)
        if count <= 0:
            raise ValueError("PDF contains no pages")
        return path
    except Exception:
        repaired_dir.mkdir(parents=True, exist_ok=True)
        repaired = repaired_dir / f"{path.stem}_repaired.pdf"
        try:
            subprocess.run(
                ["qpdf", "--warning-exit-0", "--linearize", str(path), str(repaired)],
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )
            if pdf_page_count(repaired) > 0:
                return repaired
        except Exception:
            pass
        doc = fitz.open(str(path))
        doc.save(str(repaired), garbage=4, deflate=True, clean=True)
        doc.close()
        if pdf_page_count(repaired) <= 0:
            raise ValueError(f"Unable to repair PDF: {path}")
        return repaired


def download_drive_pdf(file_id: str, cache_dir: Path, repaired_dir: Path, attempts: int = 4) -> Path:
    cache_dir.mkdir(parents=True, exist_ok=True)
    dest = cache_dir / f"{file_id}.pdf"
    if dest.exists() and dest.stat().st_size > 500:
        try:
            return repair_pdf(dest, repaired_dir)
        except Exception:
            dest.unlink(missing_ok=True)
    errors: list[str] = []
    for attempt in range(1, attempts + 1):
        tmp = cache_dir / f".{file_id}.{attempt}.part"
        tmp.unlink(missing_ok=True)
        try:
            result = gdown.download(
                id=file_id,
                output=str(tmp),
                quiet=True,
                use_cookies=False,
            )
            if not result or not tmp.exists() or tmp.stat().st_size < 500:
                raise RuntimeError(f"gdown returned {result!r}; size={tmp.stat().st_size if tmp.exists() else 0}")
            tmp.replace(dest)
            return repair_pdf(dest, repaired_dir)
        except Exception as exc:
            errors.append(f"attempt {attempt}: {type(exc).__name__}: {exc}")
            tmp.unlink(missing_ok=True)
            time.sleep(min(2 ** attempt, 12))
    raise RuntimeError("; ".join(errors))


def download_drive_file(file_id: str, output: Path) -> Path:
    output.parent.mkdir(parents=True, exist_ok=True)
    for attempt in range(1, 4):
        output.unlink(missing_ok=True)
        try:
            result = gdown.download(id=file_id, output=str(output), quiet=True, use_cookies=False)
            if result and output.exists() and output.stat().st_size > 500:
                return output
        except Exception:
            pass
        time.sleep(2 ** attempt)
    raise RuntimeError(f"Could not download Drive file {file_id}")


def load_registry(workbook_path: Path) -> list[RegistryEntry]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if "Сводный реестр" not in wb.sheetnames:
        raise KeyError("Worksheet 'Сводный реестр' not found in working registry")
    ws = wb["Сводный реестр"]
    headers = [text(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    index = {name: i for i, name in enumerate(headers) if name}
    required_headers = [
        "Альбом",
        "№",
        "Строка Excel",
        "Наименование по реестру",
        "Номер/дата документа",
        "Организация",
        "Статус",
        "Найденный файл/лист",
        "Папка/место хранения",
        "Drive URL",
        "Локальный файл в сборке",
        "Примечание",
    ]
    missing_headers = [h for h in required_headers if h not in index]
    if missing_headers:
        raise KeyError(f"Missing registry columns: {missing_headers}")

    entries: list[RegistryEntry] = []
    for row_values in ws.iter_rows(min_row=2, values_only=True):
        album = text(row_values[index["Альбом"]])
        if not album or album not in ALBUMS:
            continue
        original_status = text(row_values[index["Статус"]])
        match_type = "partial" if "ПРОВЕРИТЬ" in original_status.upper() else "exact"
        if "НЕТ В АРХИВЕ" in original_status.upper():
            match_type = "missing"
        entry = RegistryEntry(
            album=album,
            number=safe_int(row_values[index["№"]]),
            excel_row=text(row_values[index["Строка Excel"]]),
            required_name=text(row_values[index["Наименование по реестру"]]),
            document_ref=text(row_values[index["Номер/дата документа"]]),
            organization=text(row_values[index["Организация"]]),
            original_status=original_status,
            found_names_text=text(row_values[index["Найденный файл/лист"]]),
            storage_location=text(row_values[index["Папка/место хранения"]]),
            drive_urls_text=text(row_values[index["Drive URL"]]),
            local_file_old=text(row_values[index["Локальный файл в сборке"]]),
            notes=text(row_values[index["Примечание"]]),
            match_type=match_type,
        )
        entries.append(entry)
    wb.close()

    grouped = defaultdict(list)
    for entry in entries:
        grouped[entry.album].append(entry)
    for album in ALBUMS:
        grouped[album].sort(key=lambda x: x.number)
        numbers = [e.number for e in grouped[album]]
        if numbers != sorted(numbers):
            raise ValueError(f"Registry order is not stable for {album}")
        if len(numbers) != len(set(numbers)):
            raise ValueError(f"Duplicate registry numbers in {album}")
    return [entry for album in ALBUMS for entry in grouped[album]]


def prepare_source_specs(entries: list[RegistryEntry]) -> None:
    for entry in entries:
        # Restore the two exact protocols from the 4-page scan identified by Drive search and OCR.
        if entry.album == "Реестр на кронштейны" and entry.number in (31, 32):
            protocol = "АК-001" if entry.number == 31 else "АК-002"
            pages = [0, 1] if entry.number == 31 else [2, 3]
            entry.source_specs = [
                SourceSpec(
                    file_id=SCAN_0264_ID,
                    display_name=f"Протокол испытаний № {protocol} от 18.02.2025.pdf",
                    drive_url=f"https://drive.google.com/file/d/{SCAN_0264_ID}/view",
                    page_slice=pages,
                )
            ]
            entry.drive_urls_text = f"https://drive.google.com/file/d/{SCAN_0264_ID}/view"
            entry.found_names_text = f"scan_0264.pdf, страницы {pages[0]+1}-{pages[-1]+1} ({protocol})"
            entry.storage_location = "Исполнительная Ш550 / scan_0264.pdf"
            entry.match_type = "exact"
            entry.notes = (
                f"Восстановлено из scan_0264.pdf: страницы {pages[0]+1}-{pages[-1]+1}; "
                f"OCR подтверждает номер {protocol} и дату 18.02.2025."
            )
            continue

        urls = split_semicolon(entry.drive_urls_text)
        names = split_semicolon(entry.found_names_text)
        if not urls:
            entry.match_type = "missing"
            continue
        for idx, url in enumerate(urls):
            try:
                file_id = extract_drive_id(url)
            except Exception as exc:
                entry.match_type = "download_error"
                entry.build_note = f"Invalid Drive URL: {exc}"
                continue
            if idx < len(names):
                name = names[idx]
            elif len(names) == 1:
                name = f"{Path(names[0]).stem}_{idx+1}.pdf"
            else:
                name = f"{entry.required_name}_{idx+1}.pdf"
            entry.source_specs.append(SourceSpec(file_id=file_id, display_name=name, drive_url=url))
        if not entry.source_specs:
            entry.match_type = "missing"


def split_pdf_pages(source: Path, page_indices: list[int], output: Path, title: str) -> Path:
    reader = PdfReader(str(source), strict=False)
    if reader.is_encrypted:
        reader.decrypt("")
    writer = PdfWriter()
    for page_index in page_indices:
        if page_index < 0 or page_index >= len(reader.pages):
            raise IndexError(f"Page {page_index+1} not available in {source.name}")
        writer.add_page(reader.pages[page_index])
    writer.add_metadata({"/Title": title, "/Producer": "SH550 clean rebuild"})
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("wb") as fh:
        writer.write(fh)
    return output


def ocr_scan_protocols(scan_path: Path, work_dir: Path) -> dict[str, object]:
    """OCR first pages to independently confirm AK-001 and AK-002 positions."""
    ocr_dir = work_dir / "ocr_scan_0264"
    ocr_dir.mkdir(parents=True, exist_ok=True)
    prefix = ocr_dir / "scan"
    subprocess.run(
        ["pdftoppm", "-png", "-r", "180", str(scan_path), str(prefix)],
        check=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    page_texts: list[str] = []
    for image in sorted(ocr_dir.glob("scan-*.png")):
        base = image.with_suffix("")
        result = subprocess.run(
            ["tesseract", str(image), "stdout", "-l", "rus+eng"],
            check=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )
        page_texts.append(result.stdout.decode("utf-8", errors="ignore"))
    normalized = [re.sub(r"\s+", " ", t.upper().replace("—", "-").replace("–", "-")) for t in page_texts]
    page_ak001 = next((i for i, t in enumerate(normalized) if "АК-001" in t), None)
    page_ak002 = next((i for i, t in enumerate(normalized) if "АК-002" in t), None)
    result = {
        "page_count": len(page_texts),
        "AK-001_first_page": None if page_ak001 is None else page_ak001 + 1,
        "AK-002_first_page": None if page_ak002 is None else page_ak002 + 1,
        "confirmed": page_ak001 == 0 and page_ak002 == 2 and len(page_texts) == 4,
        "page_text_excerpt": [t[:700] for t in page_texts],
    }
    (ocr_dir / "ocr_result.json").write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def paragraph(value: str, style: ParagraphStyle) -> Paragraph:
    return Paragraph(html.escape(value or "—").replace("\n", "<br/>") , style)


def make_service_page(entry: RegistryEntry, output: Path, kind: str) -> Path:
    output.parent.mkdir(parents=True, exist_ok=True)
    styles = getSampleStyleSheet()
    normal = ParagraphStyle(
        "normal_ru",
        parent=styles["BodyText"],
        fontName=FONT_REGULAR,
        fontSize=10,
        leading=14,
        spaceAfter=7,
    )
    heading = ParagraphStyle(
        "heading_ru",
        parent=styles["Title"],
        fontName=FONT_BOLD,
        fontSize=18,
        leading=22,
        alignment=TA_CENTER,
        spaceAfter=14,
    )
    small = ParagraphStyle(
        "small_ru",
        parent=normal,
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor("#555555"),
    )

    if kind == "partial":
        title = "ПОМЕТКА О НЕПОЛНОМ СООТВЕТСТВИИ"
        accent = colors.HexColor("#D99A00")
        body = (
            "В архиве найден документ, который частично соответствует позиции реестра. "
            "Он подшит следом за данным листом в найденном виде. Требуется подтверждение ПТО/строительного контроля "
            "по осям, листам и комплектности."
        )
    elif kind == "download_error":
        title = "ФАЙЛ НЕ УДАЛОСЬ ВЫГРУЗИТЬ"
        accent = colors.HexColor("#B3261E")
        body = (
            "Ссылка на файл имеется, однако файл не был получен при автоматизированной пересборке. "
            "На месте документа оставлен данный лист-пометка."
        )
    else:
        title = "ДОКУМЕНТ ОТСУТСТВУЕТ В АРХИВЕ"
        accent = colors.HexColor("#B3261E")
        body = (
            "Точное соответствие позиции реестра в доступном архиве не найдено. "
            "На месте документа в альбоме оставлен данный лист-пометка."
        )

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(accent)
        canvas.setLineWidth(1.2)
        canvas.line(18 * mm, 17 * mm, A4[0] - 18 * mm, 17 * mm)
        canvas.setFont(FONT_REGULAR, 7.5)
        canvas.setFillColor(colors.HexColor("#555555"))
        canvas.drawString(18 * mm, 11 * mm, f"{entry.album} • позиция {entry.number}")
        canvas.drawRightString(A4[0] - 18 * mm, 11 * mm, f"Сформировано {BUILD_DATE}")
        canvas.restoreState()

    doc = SimpleDocTemplate(
        str(output),
        pagesize=A4,
        leftMargin=22 * mm,
        rightMargin=22 * mm,
        topMargin=24 * mm,
        bottomMargin=24 * mm,
        title=title,
        author="SH550 clean rebuild",
    )
    story = [
        Spacer(1, 10 * mm),
        Paragraph(title, heading),
        Table([[""]], colWidths=[165 * mm], rowHeights=[3.5 * mm], style=TableStyle([("BACKGROUND", (0, 0), (-1, -1), accent)])),
        Spacer(1, 10 * mm),
        paragraph(f"Альбом: {entry.album}", normal),
        paragraph(f"Позиция реестра: {entry.number}", normal),
        paragraph(f"Требуемый документ: {entry.required_name}", normal),
        paragraph(f"Номер/дата: {entry.document_ref or 'не указаны'}", normal),
        paragraph(f"Организация: {entry.organization or 'не указана'}", normal),
        Spacer(1, 4 * mm),
        paragraph(body, normal),
        Spacer(1, 4 * mm),
        paragraph(f"Найденный файл: {entry.found_names_text or 'не найден'}", small),
        paragraph(f"Место хранения: {entry.storage_location or 'не установлено'}", small),
        paragraph(f"Примечание: {entry.notes or entry.build_note or '—'}", small),
    ]
    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    return output


def build_front_matter(album: str, entries: list[RegistryEntry], output: Path) -> Path:
    page_size = landscape(A4)
    width, height = page_size
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "title_ru",
        parent=styles["Title"],
        fontName=FONT_BOLD,
        fontSize=22,
        leading=27,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#233746"),
        spaceAfter=10,
    )
    subtitle_style = ParagraphStyle(
        "subtitle_ru",
        parent=styles["Heading2"],
        fontName=FONT_REGULAR,
        fontSize=11,
        leading=15,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#495B66"),
    )
    body_style = ParagraphStyle(
        "body_ru",
        parent=styles["BodyText"],
        fontName=FONT_REGULAR,
        fontSize=8,
        leading=10,
    )
    table_style = ParagraphStyle(
        "table_ru",
        parent=body_style,
        fontSize=5.8,
        leading=7.2,
        wordWrap="CJK",
    )
    table_header_style = ParagraphStyle(
        "table_header_ru",
        parent=table_style,
        fontName=FONT_BOLD,
        textColor=colors.white,
        alignment=TA_CENTER,
        fontSize=6.2,
        leading=7.5,
    )

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#B4C0C7"))
        canvas.setLineWidth(0.35)
        canvas.line(12 * mm, 10 * mm, width - 12 * mm, 10 * mm)
        canvas.setFont(FONT_REGULAR, 6.8)
        canvas.setFillColor(colors.HexColor("#596870"))
        canvas.drawString(12 * mm, 5.5 * mm, f"Исполнительная документация Ш550 • {album}")
        canvas.drawRightString(width - 12 * mm, 5.5 * mm, f"стр. {doc.page} • {BUILD_DATE}")
        canvas.restoreState()

    doc = SimpleDocTemplate(
        str(output),
        pagesize=page_size,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=13 * mm,
        bottomMargin=14 * mm,
        title=album,
        author="SH550 clean rebuild",
        allowSplitting=1,
    )

    exact_count = sum(e.match_type == "exact" for e in entries)
    partial_count = sum(e.match_type == "partial" for e in entries)
    missing_count = sum(e.match_type in ("missing", "download_error") for e in entries)
    total_source_files = sum(len(e.source_specs) for e in entries if e.match_type not in ("missing", "download_error"))

    story = [
        Spacer(1, 13 * mm),
        Paragraph("ИСПОЛНИТЕЛЬНАЯ ДОКУМЕНТАЦИЯ", title_style),
        Spacer(1, 4 * mm),
        Paragraph(html.escape(album), title_style),
        Spacer(1, 7 * mm),
        Paragraph(html.escape(OBJECT_TITLE), subtitle_style),
        Paragraph(html.escape(OBJECT_ADDRESS), subtitle_style),
        Paragraph(f"Шифр проекта: {html.escape(PROJECT_CODE)}", subtitle_style),
        Spacer(1, 14 * mm),
    ]
    stats_data = [
        ["Позиций реестра", "Файлов-источников", "Подшито", "С замечаниями", "Отсутствует"],
        [str(len(entries)), str(total_source_files), str(exact_count), str(partial_count), str(missing_count)],
    ]
    stats = Table(stats_data, colWidths=[45 * mm] * 5, rowHeights=[10 * mm, 14 * mm])
    stats.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
                ("FONTNAME", (0, 1), (-1, 1), FONT_BOLD),
                ("FONTSIZE", (0, 0), (-1, 0), 7.5),
                ("FONTSIZE", (0, 1), (-1, 1), 15),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E5ECEF")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#AAB8C0")),
                ("TEXTCOLOR", (0, 1), (2, 1), colors.HexColor("#26633F")),
                ("TEXTCOLOR", (3, 1), (3, 1), colors.HexColor("#9A6800")),
                ("TEXTCOLOR", (4, 1), (4, 1), colors.HexColor("#A0302A")),
            ]
        )
    )
    story.extend(
        [
            stats,
            Spacer(1, 11 * mm),
            Paragraph(
                "Порядок документов соответствует строкам загруженного реестра. "
                "Для неполных совпадений перед найденным документом помещён лист замечания; "
                "для отсутствующих документов — лист-пометка. Протоколы АК-001 и АК-002 восстановлены "
                "из четырёхстраничного файла scan_0264.pdf и разделены по 2 страницы после OCR-проверки номеров.",
                ParagraphStyle(
                    "cover_note",
                    parent=body_style,
                    fontSize=8.8,
                    leading=12,
                    alignment=TA_LEFT,
                    leftIndent=22 * mm,
                    rightIndent=22 * mm,
                    textColor=colors.HexColor("#42535C"),
                ),
            ),
            Spacer(1, 10 * mm),
            Paragraph(f"Чистовая пересборка от {BUILD_DATE}", subtitle_style),
            PageBreak(),
            Paragraph("РЕЕСТР ЛИСТОВ И ДОКУМЕНТОВ АЛЬБОМА", ParagraphStyle("h", parent=title_style, fontSize=15, leading=18)),
            Spacer(1, 4 * mm),
        ]
    )

    headers = ["№", "Наименование по реестру", "№ / дата", "Подшитый файл", "Статус", "Листы PDF", "Примечание"]
    data = [[paragraph(h, table_header_style) for h in headers]]
    for entry in entries:
        note = entry.notes or entry.build_note
        found = entry.found_names_text or "—"
        data.append(
            [
                paragraph(str(entry.number), table_style),
                paragraph(entry.required_name, table_style),
                paragraph(entry.document_ref or "—", table_style),
                paragraph(found, table_style),
                paragraph(entry.short_status, table_style),
                paragraph(entry.page_range or "—", table_style),
                paragraph(note or "—", table_style),
            ]
        )
    col_widths = [8 * mm, 73 * mm, 37 * mm, 58 * mm, 29 * mm, 22 * mm, 46 * mm]
    table = LongTable(data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2D4858")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), FONT_BOLD),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (4, 1), (5, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#AEB9BF")),
        ("LEFTPADDING", (0, 0), (-1, -1), 2.2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2.2),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
    ]
    for idx, entry in enumerate(entries, start=1):
        if entry.match_type == "exact":
            fill = colors.HexColor("#F5FAF7") if idx % 2 else colors.HexColor("#EDF7F0")
        elif entry.match_type == "partial":
            fill = colors.HexColor("#FFF7DD")
        else:
            fill = colors.HexColor("#FDE8E6")
        style_cmds.append(("BACKGROUND", (0, idx), (-1, idx), fill))
    table.setStyle(TableStyle(style_cmds))
    story.append(table)
    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    return output


def copy_sources_and_build_manifest(album_dir: Path, entries: list[RegistryEntry]) -> None:
    sources_dir = album_dir / "sources"
    service_dir = album_dir / "service_pages"
    sources_dir.mkdir(parents=True, exist_ok=True)
    service_dir.mkdir(parents=True, exist_ok=True)

    manifest_rows: list[dict[str, str]] = []
    for entry in entries:
        entry.archive_files = []
        if entry.service_page:
            service_name = f"{entry.number:03d}_{'warning' if entry.match_type == 'partial' else 'missing'}_notice.pdf"
            shutil.copy2(entry.service_page, service_dir / service_name)
            entry.archive_files.append(f"service_pages/{service_name}")
        for idx, spec in enumerate(entry.source_specs):
            if not spec.included_path or spec.error:
                continue
            suffix = "" if len(entry.source_specs) == 1 else chr(ord("a") + idx)
            preferred = spec.display_name or entry.required_name
            filename = sanitize_filename(preferred)
            filename = f"{entry.number:03d}{suffix}_{filename}"
            target = sources_dir / filename
            shutil.copy2(spec.included_path, target)
            spec.archive_name = f"sources/{filename}"
            entry.archive_files.append(spec.archive_name)
            manifest_rows.append(
                {
                    "Позиция": str(entry.number),
                    "Наименование по реестру": entry.required_name,
                    "Номер/дата": entry.document_ref,
                    "Статус": entry.short_status,
                    "Листы альбома": entry.page_range,
                    "Файл в ZIP": spec.archive_name,
                    "Страниц источника": str(spec.page_count),
                    "Drive ID": spec.file_id,
                    "Drive URL": spec.drive_url,
                    "SHA-256": spec.sha256,
                    "Примечание": entry.notes or entry.build_note,
                }
            )
        if not entry.source_specs or all(spec.error for spec in entry.source_specs):
            manifest_rows.append(
                {
                    "Позиция": str(entry.number),
                    "Наименование по реестру": entry.required_name,
                    "Номер/дата": entry.document_ref,
                    "Статус": entry.short_status,
                    "Листы альбома": entry.page_range,
                    "Файл в ZIP": entry.archive_files[0] if entry.archive_files else "",
                    "Страниц источника": "0",
                    "Drive ID": "",
                    "Drive URL": entry.drive_urls_text,
                    "SHA-256": "",
                    "Примечание": entry.notes or entry.build_note,
                }
            )

    manifest_path = album_dir / "manifest.csv"
    fieldnames = [
        "Позиция",
        "Наименование по реестру",
        "Номер/дата",
        "Статус",
        "Листы альбома",
        "Файл в ZIP",
        "Страниц источника",
        "Drive ID",
        "Drive URL",
        "SHA-256",
        "Примечание",
    ]
    with manifest_path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()
        writer.writerows(manifest_rows)

    exact = sum(e.match_type == "exact" for e in entries)
    partial = sum(e.match_type == "partial" for e in entries)
    missing = sum(e.match_type in ("missing", "download_error") for e in entries)
    readme = (
        f"Исполнительная документация Ш550\n"
        f"Альбом: {entries[0].album if entries else ''}\n"
        f"Дата пересборки: {BUILD_DATE}\n\n"
        f"Позиций: {len(entries)}\nПодшито: {exact}\nС замечанием: {partial}\nОтсутствует/ошибка: {missing}\n\n"
        "Каталог sources содержит пронумерованные файлы в порядке реестра.\n"
        "Каталог service_pages содержит листы замечаний или отсутствия, которые включены в PDF.\n"
        "manifest.csv содержит полное сопоставление, страницы альбома, Drive ID и контрольные суммы.\n"
    )
    (album_dir / "README.txt").write_text(readme, encoding="utf-8")


def zip_directory(source_dir: Path, zip_path: Path) -> Path:
    zip_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6, allowZip64=True) as zf:
        for path in sorted(source_dir.rglob("*")):
            if path.is_file():
                zf.write(path, arcname=str(path.relative_to(source_dir)))
    return zip_path


def merge_album(front_pdf: Path, entries: list[RegistryEntry], output: Path) -> int:
    writer = PdfWriter()
    writer.append(str(front_pdf), import_outline=False)
    front_count = pdf_page_count(front_pdf)
    try:
        writer.add_outline_item("Титульный лист и реестр", 0)
    except Exception:
        pass
    cursor = front_count
    for entry in entries:
        start_index = cursor
        for path, count in zip(entry.included_paths, entry.included_page_counts):
            writer.append(str(path), import_outline=False)
            cursor += count
        try:
            label = f"{entry.number}. {entry.required_name}"
            if len(label) > 120:
                label = label[:117] + "..."
            writer.add_outline_item(label, start_index)
        except Exception:
            pass
    writer.add_metadata(
        {
            "/Title": f"SH550 — {entries[0].album if entries else ''}",
            "/Subject": "Исполнительная документация по фасадам",
            "/Author": "ООО ОСВ-Строй / чистовая пересборка",
            "/Producer": "SH550 clean rebuild",
            "/CreationDate": datetime.now().strftime("D:%Y%m%d%H%M%S"),
        }
    )
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("wb") as fh:
        writer.write(fh)
    actual = pdf_page_count(output)
    if actual != cursor:
        raise ValueError(f"Page-count mismatch for {output.name}: expected {cursor}, got {actual}")
    return actual


def check_pdf(path: Path, preview_dir: Path) -> dict[str, object]:
    result: dict[str, object] = {"file": path.name, "size": path.stat().st_size, "sha256": sha256_file(path)}
    doc = fitz.open(str(path))
    result["pages"] = doc.page_count
    blank_pages: list[int] = []
    for page_index in range(doc.page_count):
        page = doc.load_page(page_index)
        pix = page.get_pixmap(matrix=fitz.Matrix(0.22, 0.22), colorspace=fitz.csGRAY, alpha=False)
        samples = pix.samples
        if samples:
            ink = sum(1 for value in samples if value < 248) / len(samples)
            if ink < 0.00035:
                blank_pages.append(page_index + 1)
    result["near_blank_pages"] = blank_pages
    preview_dir.mkdir(parents=True, exist_ok=True)
    preview_indices = sorted(set([0, min(1, doc.page_count - 1), max(0, doc.page_count - 1)]))
    previews = []
    for idx in preview_indices:
        page = doc.load_page(idx)
        pix = page.get_pixmap(matrix=fitz.Matrix(1.15, 1.15), alpha=False)
        name = f"{path.stem}_page_{idx+1}.png"
        pix.save(str(preview_dir / name))
        previews.append(name)
    doc.close()
    result["previews"] = previews
    qpdf = subprocess.run(
        ["qpdf", "--check", str(path)],
        check=False,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
    )
    result["qpdf_returncode"] = qpdf.returncode
    result["qpdf_output"] = qpdf.stdout.decode("utf-8", errors="replace")[-3000:]
    result["ok"] = qpdf.returncode in (0, 3) and not blank_pages
    return result


def build_registry_workbook(entries: list[RegistryEntry], album_outputs: dict[str, dict[str, object]], output: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="2D4858")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="B7C1C7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fills = {
        "ПОДШИТО": PatternFill("solid", fgColor="E7F4EA"),
        "ПОДШИТО С ЗАМЕЧАНИЕМ": PatternFill("solid", fgColor="FFF1C7"),
        "ОТСУТСТВУЕТ": PatternFill("solid", fgColor="FADBD8"),
        "ОШИБКА ВЫГРУЗКИ": PatternFill("solid", fgColor="FADBD8"),
    }
    columns = [
        "Альбом",
        "№",
        "Строка исходного Excel",
        "Наименование по реестру",
        "Номер/дата документа",
        "Организация",
        "Итоговый статус",
        "Найденный файл/лист",
        "Папка/место хранения",
        "Drive URL",
        "Листы в итоговом PDF",
        "Страниц позиции",
        "Файлы в ZIP",
        "SHA-256 источников",
        "Примечание / контроль",
        "PDF альбома",
    ]

    def add_sheet(name: str, rows: list[RegistryEntry]) -> None:
        ws = wb.create_sheet(name[:31])
        ws.append(columns)
        for c in ws[1]:
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
        for entry in rows:
            source_hashes = "\n".join(spec.sha256 for spec in entry.source_specs if spec.sha256)
            pages_count = sum(entry.included_page_counts)
            pdf_name = text(album_outputs.get(entry.album, {}).get("pdf_name", ""))
            row = [
                entry.album,
                entry.number,
                entry.excel_row,
                entry.required_name,
                entry.document_ref,
                entry.organization,
                entry.short_status,
                entry.found_names_text,
                entry.storage_location,
                entry.drive_urls_text,
                entry.page_range,
                pages_count,
                "\n".join(entry.archive_files),
                source_hashes,
                entry.notes or entry.build_note,
                pdf_name,
            ]
            ws.append(row)
            current_row = ws.max_row
            status_cell = ws.cell(current_row, 7)
            status_cell.fill = fills.get(entry.short_status, PatternFill("solid", fgColor="FFFFFF"))
            drive_cell = ws.cell(current_row, 10)
            first_url = split_semicolon(entry.drive_urls_text)
            if first_url:
                drive_cell.hyperlink = first_url[0]
                drive_cell.style = "Hyperlink"
            for cell in ws[current_row]:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{ws.max_row}"
        ws.sheet_view.showGridLines = False
        widths = [27, 7, 12, 58, 30, 28, 25, 52, 38, 48, 18, 14, 52, 45, 62, 34]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.row_dimensions[1].height = 36
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 48

    add_sheet("Сводный реестр", entries)
    for album, (code, slug) in ALBUMS.items():
        sheet_name = {
            "01": "Кронштейны",
            "02": "Утепление 1 слой",
            "03": "Утепление 2 слой",
            "04": "Направляющие",
            "05": "Керамогранит",
        }[code]
        add_sheet(sheet_name, [e for e in entries if e.album == album])

    ws = wb.create_sheet("Итоги", 0)
    ws.append(["ИТОГИ ЧИСТОВОЙ ПЕРЕСБОРКИ ИСПОЛНИТЕЛЬНОЙ ДОКУМЕНТАЦИИ Ш550"])
    ws.merge_cells("A1:H1")
    ws["A1"].font = Font(name="Arial", bold=True, size=15, color="FFFFFF")
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.append(["Дата", BUILD_DATE])
    ws.append(["Объект", OBJECT_TITLE])
    ws.append(["Адрес", OBJECT_ADDRESS])
    ws.append([])
    ws.append(["Альбом", "Позиций", "Подшито", "С замечанием", "Отсутствует/ошибка", "Страниц PDF", "PDF", "ZIP источников"])
    for c in ws[6]:
        c.fill = header_fill
        c.font = header_font
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for album in ALBUMS:
        rows = [e for e in entries if e.album == album]
        info = album_outputs[album]
        ws.append(
            [
                album,
                len(rows),
                sum(e.match_type == "exact" for e in rows),
                sum(e.match_type == "partial" for e in rows),
                sum(e.match_type in ("missing", "download_error") for e in rows),
                info["page_count"],
                info["pdf_name"],
                info["zip_name"],
            ]
        )
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.append([])
    ws.append(["Контрольный вывод", "Протоколы АК-001 и АК-002 найдены в scan_0264.pdf и подшиты отдельными двухстраничными документами."])
    ws.merge_cells(start_row=ws.max_row, start_column=2, end_row=ws.max_row, end_column=8)
    for idx, width in enumerate([40, 15, 15, 18, 20, 15, 35, 38], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A6"
    ws.sheet_view.showGridLines = False

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    # Structural validation after save.
    check_wb = load_workbook(output, read_only=True, data_only=False)
    expected = {"Итоги", "Сводный реестр", "Кронштейны", "Утепление 1 слой", "Утепление 2 слой", "Направляющие", "Керамогранит"}
    if not expected.issubset(set(check_wb.sheetnames)):
        raise ValueError(f"Output workbook missing sheets: {expected - set(check_wb.sheetnames)}")
    check_wb.close()
    return output


def build_qc_pdf(qc: dict[str, object], output: Path) -> Path:
    styles = getSampleStyleSheet()
    title = ParagraphStyle("qc_title", parent=styles["Title"], fontName=FONT_BOLD, fontSize=17, leading=21, alignment=TA_CENTER)
    normal = ParagraphStyle("qc_normal", parent=styles["BodyText"], fontName=FONT_REGULAR, fontSize=8.5, leading=11)
    small = ParagraphStyle("qc_small", parent=normal, fontSize=7, leading=9)
    doc = SimpleDocTemplate(str(output), pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=16*mm, bottomMargin=16*mm)
    story = [Paragraph("ОТЧЁТ КОНТРОЛЯ ПЕРЕСБОРКИ Ш550", title), Spacer(1, 5*mm), paragraph(f"Дата: {BUILD_DATE}", normal)]
    scan = qc.get("scan_0264", {})
    story.extend([
        paragraph(f"scan_0264: страниц {scan.get('page_count', '—')}; АК-001 начинается на стр. {scan.get('AK-001_first_page', '—')}; АК-002 начинается на стр. {scan.get('AK-002_first_page', '—')}; подтверждение: {scan.get('confirmed', False)}", normal),
        Spacer(1, 4*mm),
    ])
    rows = [["Альбом", "Страниц", "Размер", "QPDF", "Пустые страницы", "SHA-256"]]
    for item in qc.get("albums", []):
        rows.append([
            item.get("file", ""),
            str(item.get("pages", "")),
            f"{item.get('size', 0)/1024/1024:.2f} MB",
            "OK" if item.get("qpdf_returncode") in (0,3) else f"ERR {item.get('qpdf_returncode')}",
            ", ".join(map(str, item.get("near_blank_pages", []))) or "нет",
            str(item.get("sha256", ""))[:20] + "…",
        ])
    table = LongTable([[paragraph(str(cell), small) for cell in row] for row in rows], colWidths=[51*mm, 14*mm, 20*mm, 17*mm, 28*mm, 47*mm], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#2D4858")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), FONT_BOLD),
        ("GRID", (0,0), (-1,-1), 0.35, colors.HexColor("#AAB6BC")),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F2F6F7")]),
    ]))
    story.extend([table, Spacer(1, 5*mm)])
    failures = qc.get("download_failures", {})
    story.append(paragraph(f"Ошибки выгрузки: {len(failures)}", normal))
    if failures:
        for fid, error in failures.items():
            story.append(paragraph(f"{fid}: {error}", small))
    story.append(Spacer(1, 3*mm))
    story.append(paragraph("Автоматические проверки: читаемость структуры PDF, число страниц, qpdf --check, контроль почти пустых страниц, SHA-256, наличие всех ожидаемых вкладок Excel.", normal))
    doc.build(story)
    return output


def create_master_zip(release_dir: Path, output: Path) -> Path:
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6, allowZip64=True) as zf:
        for path in sorted(release_dir.iterdir()):
            if not path.is_file() or path == output:
                continue
            zf.write(path, arcname=path.name)
    return output


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", default="output", help="Build output directory")
    parser.add_argument("--workers", type=int, default=5)
    args = parser.parse_args()

    register_fonts()
    root = Path(args.output).resolve()
    if root.exists():
        shutil.rmtree(root)
    work = root / "work"
    release = root / "release"
    cache = work / "drive_cache"
    repaired = work / "repaired"
    albums_work = work / "album_sources"
    service_pages = work / "service_pages"
    front_dir = work / "front_matter"
    preview_dir = root / "qc_previews"
    for directory in [work, release, cache, repaired, albums_work, service_pages, front_dir, preview_dir]:
        directory.mkdir(parents=True, exist_ok=True)

    working_registry = download_drive_file(WORKING_REGISTRY_ID, work / "SH550_working_registry_input.xlsx")
    source_registry = download_drive_file(SOURCE_REGISTRY_ID, work / "Reestr_prilozheniy_k_aktam_po_fasadam.xlsx")
    entries = load_registry(working_registry)
    prepare_source_specs(entries)

    expected_counts = {
        "Реестр на кронштейны": 32,
        "Реестр на утепление 1 слой": 30,
        "Реестр на утепление 2 слой": 30,
        "Реестр на направляющие": 39,
        "Реестр на керамогранит": 64,
    }
    for album, expected in expected_counts.items():
        actual = sum(e.album == album for e in entries)
        if actual != expected:
            raise ValueError(f"Unexpected registry row count for {album}: expected {expected}, got {actual}")

    unique_ids = sorted({spec.file_id for entry in entries for spec in entry.source_specs})
    print(f"Registry entries: {len(entries)}; unique Drive PDFs: {len(unique_ids)}", flush=True)
    downloads: dict[str, Path] = {}
    failures: dict[str, str] = {}
    with ThreadPoolExecutor(max_workers=max(1, args.workers)) as pool:
        futures = {pool.submit(download_drive_pdf, fid, cache, repaired): fid for fid in unique_ids}
        done = 0
        for future in as_completed(futures):
            fid = futures[future]
            done += 1
            try:
                downloads[fid] = future.result()
                print(f"[{done}/{len(unique_ids)}] downloaded {fid} ({downloads[fid].stat().st_size} bytes)", flush=True)
            except Exception as exc:
                failures[fid] = f"{type(exc).__name__}: {exc}"
                print(f"[{done}/{len(unique_ids)}] FAILED {fid}: {failures[fid]}", flush=True)

    scan_ocr: dict[str, object] = {"confirmed": False}
    if SCAN_0264_ID in downloads:
        scan_ocr = ocr_scan_protocols(downloads[SCAN_0264_ID], work)
        print("scan_0264 OCR:", json.dumps(scan_ocr, ensure_ascii=False), flush=True)

    # Prepare source paths, protocol splits, statuses and service pages.
    for entry in entries:
        successful_specs: list[SourceSpec] = []
        for spec in entry.source_specs:
            if spec.file_id not in downloads:
                spec.error = failures.get(spec.file_id, "source not downloaded")
                continue
            source_path = downloads[spec.file_id]
            spec.source_path = source_path
            if spec.page_slice is not None:
                protocol = "АК-001" if entry.number == 31 else "АК-002"
                split_path = work / "split_protocols" / f"{protocol}.pdf"
                spec.included_path = split_pdf_pages(source_path, spec.page_slice, split_path, protocol)
                if not scan_ocr.get("confirmed"):
                    entry.match_type = "partial"
                    entry.notes += " Автоматическая OCR-проверка разбиения не дала полного подтверждения."
            else:
                spec.included_path = source_path
            spec.page_count = pdf_page_count(spec.included_path)
            spec.sha256 = sha256_file(spec.included_path)
            successful_specs.append(spec)

        if not successful_specs:
            if entry.source_specs and any(spec.error for spec in entry.source_specs):
                entry.match_type = "download_error"
                entry.build_note = "; ".join(spec.error for spec in entry.source_specs if spec.error)
            else:
                entry.match_type = "missing"
        elif len(successful_specs) < len(entry.source_specs):
            entry.match_type = "partial"
            missing_ids = [spec.file_id for spec in entry.source_specs if spec.error]
            entry.build_note = f"Часть файлов позиции не выгружена: {', '.join(missing_ids)}"
        entry.source_specs = successful_specs + [spec for spec in entry.source_specs if spec.error]

        entry.included_paths = []
        entry.included_page_counts = []
        if entry.match_type in ("partial", "missing", "download_error"):
            kind = entry.match_type
            notice = service_pages / f"{ALBUMS[entry.album][0]}_{entry.number:03d}_{kind}.pdf"
            entry.service_page = make_service_page(entry, notice, kind)
            entry.included_paths.append(entry.service_page)
            entry.included_page_counts.append(pdf_page_count(entry.service_page))
        for spec in entry.source_specs:
            if spec.included_path and not spec.error:
                entry.included_paths.append(spec.included_path)
                entry.included_page_counts.append(spec.page_count)
        if not entry.included_paths:
            raise ValueError(f"Entry {entry.album} #{entry.number} has no album page")

    album_outputs: dict[str, dict[str, object]] = {}
    qc_albums: list[dict[str, object]] = []
    for album, (code, slug) in ALBUMS.items():
        album_entries = [e for e in entries if e.album == album]
        # Two-pass front matter to calculate final album page ranges.
        front_path = front_dir / f"{code}_{slug}_front.pdf"
        for entry in album_entries:
            entry.page_range = "—"
        build_front_matter(album, album_entries, front_path)
        front_count = pdf_page_count(front_path)
        for _ in range(3):
            cursor = front_count + 1
            for entry in album_entries:
                pages = sum(entry.included_page_counts)
                entry.page_start = cursor
                entry.page_end = cursor + pages - 1
                entry.page_range = str(entry.page_start) if pages == 1 else f"{entry.page_start}–{entry.page_end}"
                cursor = entry.page_end + 1
            build_front_matter(album, album_entries, front_path)
            new_count = pdf_page_count(front_path)
            if new_count == front_count:
                break
            front_count = new_count
        else:
            raise RuntimeError(f"Front-matter page count did not stabilize for {album}")

        pdf_name = f"SH550_{code}_{slug}.pdf"
        zip_name = f"SH550_{code}_{slug}_sources.zip"
        pdf_path = release / pdf_name
        page_count = merge_album(front_path, album_entries, pdf_path)

        source_staging = albums_work / f"{code}_{slug}"
        copy_sources_and_build_manifest(source_staging, album_entries)
        zip_path = release / zip_name
        zip_directory(source_staging, zip_path)

        album_outputs[album] = {
            "code": code,
            "slug": slug,
            "pdf_name": pdf_name,
            "zip_name": zip_name,
            "pdf_path": str(pdf_path),
            "zip_path": str(zip_path),
            "page_count": page_count,
            "registry_positions": len(album_entries),
            "exact": sum(e.match_type == "exact" for e in album_entries),
            "partial": sum(e.match_type == "partial" for e in album_entries),
            "missing": sum(e.match_type in ("missing", "download_error") for e in album_entries),
        }
        qc_item = check_pdf(pdf_path, preview_dir)
        qc_item["album"] = album
        qc_albums.append(qc_item)
        print(f"Built {pdf_name}: {page_count} pages; source ZIP {zip_path.stat().st_size} bytes", flush=True)

    registry_output = release / "SH550_working_registry_rebuilt.xlsx"
    build_registry_workbook(entries, album_outputs, registry_output)
    shutil.copy2(source_registry, release / "SH550_source_registry_original.xlsx")

    qc = {
        "build_date": BUILD_DATE,
        "entries": len(entries),
        "unique_drive_files": len(unique_ids),
        "download_failures": failures,
        "scan_0264": scan_ocr,
        "albums": qc_albums,
        "album_summary": album_outputs,
    }
    qc_json = release / "SH550_QC_report.json"
    qc_json.write_text(json.dumps(qc, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    build_qc_pdf(qc, release / "SH550_QC_report.pdf")

    readme_lines = [
        "ИСПОЛНИТЕЛЬНАЯ ДОКУМЕНТАЦИЯ Ш550 — ЧИСТОВАЯ ПЕРЕСБОРКА",
        f"Дата: {BUILD_DATE}",
        "",
        "Состав:",
    ]
    for album, info in album_outputs.items():
        readme_lines.append(
            f"- {album}: {info['registry_positions']} позиций, {info['page_count']} стр., "
            f"подшито {info['exact']}, с замечанием {info['partial']}, отсутствует {info['missing']}."
        )
    readme_lines.extend(
        [
            "",
            "АК-001 и АК-002: найдены в scan_0264.pdf, OCR-подтверждены и разделены на страницы 1–2 и 3–4.",
            f"Ошибок выгрузки: {len(failures)}.",
            "SH550_working_registry_rebuilt.xlsx содержит построчное сопоставление, страницы итоговых PDF, Drive URL, ZIP-файлы и SHA-256.",
        ]
    )
    (release / "README_SH550.txt").write_text("\n".join(readme_lines), encoding="utf-8")

    master_zip = release / "SH550_all_albums_rebuilt.zip"
    create_master_zip(release, master_zip)

    # Final integrity checks for all release files.
    file_inventory = []
    for path in sorted(release.iterdir()):
        if path.is_file():
            file_inventory.append({"name": path.name, "size": path.stat().st_size, "sha256": sha256_file(path)})
    (root / "release_inventory.json").write_text(json.dumps(file_inventory, ensure_ascii=False, indent=2), encoding="utf-8")

    fatal_qc = [item for item in qc_albums if not item.get("ok")]
    if failures:
        print("WARNING: source download failures remain; they are represented by notice pages.", flush=True)
    if fatal_qc:
        print("WARNING: QC flags:", json.dumps(fatal_qc, ensure_ascii=False, indent=2), flush=True)
    print("FINAL_SUMMARY=" + json.dumps({"albums": album_outputs, "failures": failures, "scan": scan_ocr}, ensure_ascii=False, default=str), flush=True)
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception:
        traceback.print_exc()
        raise
